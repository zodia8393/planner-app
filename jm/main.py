"""
Work Planner - Professional task & schedule management
FastAPI + Jinja2 + HTMX + Tailwind CSS + SQLite
"""

import os
import sys
from pathlib import Path as _P
# Add parent directory to sys.path so we can import from common
_app_dir = _P(__file__).resolve().parent
sys.path.insert(0, str(_app_dir.parent))
sys.path.insert(0, str(_app_dir))
_env_path = _P(__file__).resolve().parent.parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())
import json
import uuid
import sqlite3
import calendar as cal_mod
import asyncio
import logging
import shutil
import zipfile
import io
from datetime import datetime, date, timedelta
from pathlib import Path
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, Request, Form, Query, HTTPException, Response, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware

import uvicorn

# ── Common module imports ──
from common.utils import fix_mojibake, clamp_priority, validate_date_str, validate_datetime_str, clamp_text
from common.filters import register_filters, render_error_page as _render_error_page, render_worklog_images
from common.middleware import EventBus, CSRFMiddleware, SyncBroadcastMiddleware, patch_formparser_utf8
from common.db import get_db as _common_get_db
from common.image import MAGIC_BYTES, _check_image_magic
from common.gcal import (
    GCAL_CLIENT_ID, GCAL_CLIENT_SECRET, GCAL_SCOPES, GCAL_AUTH_URL,
    GCAL_TOKEN_URL, GCAL_API_BASE,
    gcal_redirect_uri as _gcal_redirect_uri,
    gcal_refresh_token as _common_gcal_refresh_token,
    gcal_fetch_events as _common_gcal_fetch_events,
    gcal_push_event as _common_gcal_push_event,
    gcal_update_event as _common_gcal_update_event,
    gcal_delete_event as _common_gcal_delete_event,
)
from common.excel import parse_excel_with_merges as _parse_excel_with_merges, infer_field_type as _infer_field_type
from common.holidays import KOREAN_HOLIDAYS, get_holidays_for_month


# ── Starlette FormParser latin-1 -> utf-8 patch ──
patch_formparser_utf8()


# ── SSE EventBus (from common) ──
event_bus = EventBus()


# ── Input validation helpers (imported from common.utils) ──




# ── Path setup ──
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "work.db"
(BASE_DIR / "data").mkdir(parents=True, exist_ok=True)


# ── FastAPI app ──
@asynccontextmanager
async def lifespan(app):
    init_db()
    yield

app = FastAPI(title="Work Planner", docs_url=None, redoc_url=None, lifespan=lifespan)


access_logger = logging.getLogger("access")
access_logger.setLevel(logging.INFO)
if not access_logger.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(message)s"))
    access_logger.addHandler(_h)


class AccessLogMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/static") or request.url.path == "/sse":
            return response
        client_ip = request.headers.get("fly-client-ip",
                    request.headers.get("x-forwarded-for", "").split(",")[0].strip()
                    or (request.client.host if request.client else "unknown"))
        access_logger.info(
            f"[ACCESS] {client_ip} {request.method} {request.url.path} → {response.status_code}"
        )
        return response


# CSRFMiddleware and SyncBroadcastMiddleware imported from common.middleware

app.add_middleware(SyncBroadcastMiddleware, event_bus=event_bus,
                   skip_paths=("/worklogs/upload-image",),
                   skip_prefixes=("/files/",))
app.add_middleware(CSRFMiddleware)
app.add_middleware(AccessLogMiddleware)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
BG_DIR = BASE_DIR / "data" / "backgrounds"
BG_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/backgrounds", StaticFiles(directory=str(BG_DIR)), name="backgrounds")
WORKLOG_IMG_DIR = BASE_DIR / "data" / "worklog_images"
WORKLOG_IMG_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/worklog-images", StaticFiles(directory=str(WORKLOG_IMG_DIR)), name="worklog_images")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def render(request: Request, name: str, context: dict = None):
    """TemplateResponse wrapper that injects the single profile."""
    ctx = context or {}
    try:
        with get_db() as conn:
            row = conn.execute("SELECT * FROM work_profiles WHERE id=1").fetchone()
            if row:
                ctx["active_profile"] = dict(row)
                ctx["active_profile_id"] = 1
            else:
                ctx["active_profile"] = {"id": 1, "name": "정미", "emoji": "💼", "role": ""}
                ctx["active_profile_id"] = 1
            # Inject background setting
            bg_type = get_user_setting(conn, 1, "bg_type", "none")
            bg_preset = get_user_setting(conn, 1, "bg_preset", "")
            bg_image = get_user_setting(conn, 1, "bg_image", "")
            bg_opacity = get_user_setting(conn, 1, "bg_opacity", "0.7")
            ctx["bg_setting"] = {
                "type": bg_type,
                "preset": bg_preset,
                "image": bg_image,
                "opacity": float(bg_opacity),
            }
    except Exception:
        ctx["active_profile"] = {"id": 1, "name": "정미", "emoji": "💼", "role": ""}
        ctx["active_profile_id"] = 1
        ctx["bg_setting"] = {"type": "none", "preset": "", "image": "", "opacity": 0.7}
    ctx["needs_pin_setup"] = False
    ctx.setdefault("today", date.today())
    return templates.TemplateResponse(request, name, ctx)


# ── SSE endpoint ──
@app.get("/sse")
async def sse_stream(request: Request):
    sid, queue = event_bus.subscribe()

    async def generate():
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    page = await asyncio.wait_for(queue.get(), timeout=30)
                    yield f"event: sync\ndata: {page}\n\n"
                except asyncio.TimeoutError:
                    yield ": keepalive\n\n"
        finally:
            event_bus.unsubscribe(sid)

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Jinja2 filters (from common.filters) ──
# _render_error_page imported as render_error_page from common.filters


@app.exception_handler(404)
async def not_found_handler(request: Request, exc):
    return HTMLResponse(_render_error_page(404, "페이지를 찾을 수 없습니다"), status_code=404)


@app.exception_handler(500)
async def server_error_handler(request: Request, exc):
    return HTMLResponse(_render_error_page(500, "서버 오류가 발생했습니다"), status_code=500)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    if isinstance(exc, HTTPException):
        if exc.status_code == 404:
            return HTMLResponse(_render_error_page(404, "페이지를 찾을 수 없습니다"), status_code=404)
        raise exc
    return HTMLResponse(_render_error_page(500, "서버 오류가 발생했습니다"), status_code=500)

register_filters(templates)


# ── DB management (wrapper around common.db.get_db) ──
def get_db():
    return _common_get_db(DB_PATH)


def init_db():
    """Initialize DB schema."""
    with get_db() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS work_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            role TEXT DEFAULT '',
            emoji TEXT DEFAULT '💼',
            pin TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            pin_hash TEXT DEFAULT '',
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        );
        INSERT OR IGNORE INTO app_settings (id) VALUES (1);

        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            color TEXT NOT NULL DEFAULT '#6366f1',
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS todos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            priority INTEGER DEFAULT 2,
            category_id INTEGER,
            due_date TEXT,
            completed INTEGER DEFAULT 0,
            completed_at TEXT,
            sort_order INTEGER DEFAULT 0,
            repeat_type TEXT DEFAULT 'none',
            tags TEXT DEFAULT '[]',
            assignee TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS subtasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            todo_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            completed INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (todo_id) REFERENCES todos(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            start_time TEXT NOT NULL,
            end_time TEXT,
            all_day INTEGER DEFAULT 0,
            category_id INTEGER,
            color TEXT DEFAULT '#6366f1',
            memo TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE SET NULL
        );

        CREATE TABLE IF NOT EXISTS memos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            author TEXT NOT NULL DEFAULT '',
            content TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS ddays (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            title TEXT NOT NULL,
            target_date TEXT NOT NULL,
            icon TEXT DEFAULT '🎯',
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS work_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            log_date TEXT NOT NULL DEFAULT (date('now', 'localtime')),
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            hours REAL DEFAULT 0,
            category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL,
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS notices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            content TEXT DEFAULT '',
            priority INTEGER DEFAULT 0,
            pinned INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS form_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            fields TEXT NOT NULL DEFAULT '[]',
            emoji TEXT DEFAULT '📝',
            color TEXT DEFAULT '#6366f1',
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS form_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            template_id INTEGER NOT NULL,
            profile_id INTEGER NOT NULL,
            entry_date TEXT NOT NULL DEFAULT (date('now', 'localtime')),
            values_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT DEFAULT (datetime('now', 'localtime')),
            updated_at TEXT DEFAULT (datetime('now', 'localtime')),
            FOREIGN KEY (template_id) REFERENCES form_templates(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS gcal_tokens (
            profile_id INTEGER PRIMARY KEY,
            access_token TEXT NOT NULL,
            refresh_token TEXT NOT NULL,
            token_expiry TEXT NOT NULL,
            calendar_id TEXT DEFAULT 'primary',
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS user_settings (
            profile_id TEXT NOT NULL,
            key TEXT NOT NULL,
            value TEXT,
            PRIMARY KEY(profile_id, key)
        );

        CREATE TABLE IF NOT EXISTS links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER DEFAULT 1,
            title TEXT NOT NULL,
            url TEXT NOT NULL,
            category TEXT DEFAULT '',
            description TEXT DEFAULT '',
            favicon TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS time_budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            category_id INTEGER REFERENCES categories(id) ON DELETE CASCADE,
            weekly_hours REAL NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS todo_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            name TEXT NOT NULL,
            items_json TEXT NOT NULL DEFAULT '[]',
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS automation_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id INTEGER NOT NULL DEFAULT 0,
            name TEXT NOT NULL,
            trigger_type TEXT NOT NULL DEFAULT 'weekly',
            trigger_config TEXT NOT NULL DEFAULT '{}',
            action_type TEXT NOT NULL DEFAULT 'create_todo',
            action_config TEXT NOT NULL DEFAULT '{}',
            enabled INTEGER DEFAULT 1,
            last_run TEXT,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            changes_json TEXT DEFAULT '{}',
            created_at TEXT DEFAULT (datetime('now','localtime')),
            profile_id TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_todos_due ON todos(due_date);
        CREATE INDEX IF NOT EXISTS idx_todos_completed ON todos(completed);
        CREATE INDEX IF NOT EXISTS idx_events_start ON events(start_time);
        CREATE INDEX IF NOT EXISTS idx_work_logs_date ON work_logs(log_date);
        CREATE INDEX IF NOT EXISTS idx_form_entries_template ON form_entries(template_id);
        CREATE INDEX IF NOT EXISTS idx_form_entries_date ON form_entries(entry_date);
        CREATE INDEX IF NOT EXISTS idx_audit_entity ON audit_log(entity_type, entity_id);
        """)

        for tbl in ("work_profiles", ):
            try:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN role TEXT DEFAULT ''")
            except sqlite3.OperationalError:
                pass
        for tbl in ("todos", "events", "memos", "ddays"):
            try:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN profile_id INTEGER NOT NULL DEFAULT 0")
            except sqlite3.OperationalError:
                pass

        # Memo title + category migration
        try:
            conn.execute("ALTER TABLE memos ADD COLUMN title TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            conn.execute("ALTER TABLE memos ADD COLUMN category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL")
        except sqlite3.OperationalError:
            pass

        # Migration: add gcal_event_id to events
        ev_cols = [r[1] for r in conn.execute("PRAGMA table_info(events)").fetchall()]
        if "gcal_event_id" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN gcal_event_id TEXT DEFAULT ''")

        # Migration: add emoji/color to form_templates
        ft_cols = [r[1] for r in conn.execute("PRAGMA table_info(form_templates)").fetchall()]
        if "emoji" not in ft_cols:
            conn.execute("ALTER TABLE form_templates ADD COLUMN emoji TEXT DEFAULT '📝'")
        if "color" not in ft_cols:
            conn.execute("ALTER TABLE form_templates ADD COLUMN color TEXT DEFAULT '#6366f1'")

        # Migration: add recurrence columns to todos and events
        todo_cols = [r[1] for r in conn.execute("PRAGMA table_info(todos)").fetchall()]
        if "recurrence_end" not in todo_cols:
            conn.execute("ALTER TABLE todos ADD COLUMN recurrence_end TEXT DEFAULT ''")
        if "recurrence" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN recurrence TEXT DEFAULT ''")
        if "recurrence_end" not in ev_cols:
            conn.execute("ALTER TABLE events ADD COLUMN recurrence_end TEXT DEFAULT ''")

        # Migration: add energy_level column to todos (1=Low, 2=Medium, 3=High)
        if "energy_level" not in todo_cols:
            conn.execute("ALTER TABLE todos ADD COLUMN energy_level INTEGER DEFAULT 2")

        existing = conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        if existing == 0:
            conn.executemany(
                "INSERT INTO categories (name, color, sort_order) VALUES (?, ?, ?)",
                [("업무", "#6366f1", 0), ("회의", "#8b5cf6", 1), ("개인", "#10b981", 2), ("기타", "#f59e0b", 3)],
            )

        existing_profiles = conn.execute("SELECT COUNT(*) FROM work_profiles").fetchone()[0]
        if existing_profiles == 0:
            conn.execute(
                "INSERT INTO work_profiles (name, emoji, role) VALUES (?, ?, ?)",
                ("정미", "💼", ""),
            )

    _seed_form_presets()


def _seed_form_presets():
    """Seed Korean admin form presets if not already present."""
    presets = [
        (
            "품의서", "기안/승인 요청 양식", "📋", "#6366f1",
            [
                {"name": "title", "label": "제목", "type": "text", "required": True},
                {"name": "draft_date", "label": "기안일", "type": "date", "required": True},
                {"name": "drafter", "label": "기안자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "amount", "label": "금액", "type": "number", "required": False},
                {"name": "purpose", "label": "목적", "type": "textarea", "required": True},
                {"name": "details", "label": "상세내역", "type": "textarea", "required": False},
                {"name": "attachment_memo", "label": "첨부파일 메모", "type": "text", "required": False},
                {"name": "approval_status", "label": "결재상태", "type": "select", "required": True, "options": ["대기", "승인", "반려"]},
            ],
        ),
        (
            "회의록", "회의 기록 양식", "🗓️", "#8b5cf6",
            [
                {"name": "meeting_name", "label": "회의명", "type": "text", "required": True},
                {"name": "meeting_datetime", "label": "일시", "type": "datetime", "required": True},
                {"name": "location", "label": "장소", "type": "text", "required": False},
                {"name": "attendees", "label": "참석자", "type": "textarea", "required": True},
                {"name": "agenda", "label": "안건", "type": "textarea", "required": True},
                {"name": "discussion", "label": "논의내용", "type": "textarea", "required": False},
                {"name": "decisions", "label": "결정사항", "type": "textarea", "required": False},
                {"name": "follow_up", "label": "후속조치", "type": "textarea", "required": False},
            ],
        ),
        (
            "업무일지", "일일 업무 보고 양식", "📝", "#10b981",
            [
                {"name": "report_date", "label": "날짜", "type": "date", "required": True},
                {"name": "author", "label": "작성자", "type": "text", "required": True},
                {"name": "today_work", "label": "오늘 업무", "type": "textarea", "required": True},
                {"name": "tomorrow_plan", "label": "내일 계획", "type": "textarea", "required": False},
                {"name": "issues", "label": "이슈/건의", "type": "textarea", "required": False},
                {"name": "work_hours", "label": "근무시간", "type": "number", "required": False},
            ],
        ),
        (
            "출장보고서", "출장 결과 보고 양식", "✈️", "#0ea5e9",
            [
                {"name": "title", "label": "출장 제목", "type": "text", "required": True},
                {"name": "traveler", "label": "출장자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "start_date", "label": "출장 시작일", "type": "date", "required": True},
                {"name": "end_date", "label": "출장 종료일", "type": "date", "required": True},
                {"name": "destination", "label": "출장지", "type": "text", "required": True},
                {"name": "purpose", "label": "출장 목적", "type": "textarea", "required": True},
                {"name": "result", "label": "출장 성과", "type": "textarea", "required": True},
                {"name": "expenses", "label": "지출 내역", "type": "textarea", "required": False},
                {"name": "total_cost", "label": "총 비용(원)", "type": "number", "required": False},
                {"name": "follow_up", "label": "후속 조치", "type": "textarea", "required": False},
            ],
        ),
        (
            "지출결의서", "경비 지출 승인 양식", "💰", "#f59e0b",
            [
                {"name": "title", "label": "지출 건명", "type": "text", "required": True},
                {"name": "request_date", "label": "신청일", "type": "date", "required": True},
                {"name": "requester", "label": "신청자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "category", "label": "지출 구분", "type": "select", "required": True, "options": ["업무추진비", "여비교통비", "소모품비", "교육훈련비", "통신비", "기타"]},
                {"name": "amount", "label": "금액(원)", "type": "number", "required": True},
                {"name": "payment_method", "label": "결제수단", "type": "select", "required": False, "options": ["법인카드", "개인카드(추후정산)", "현금", "계좌이체"]},
                {"name": "details", "label": "지출 상세", "type": "textarea", "required": True},
                {"name": "receipt_note", "label": "영수증 비고", "type": "text", "required": False},
                {"name": "approval_status", "label": "결재상태", "type": "select", "required": True, "options": ["대기", "승인", "반려"]},
            ],
        ),
        (
            "휴가신청서", "연차/반차/특별휴가 신청", "🏖️", "#14b8a6",
            [
                {"name": "applicant", "label": "신청자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "leave_type", "label": "휴가 종류", "type": "select", "required": True, "options": ["연차", "오전반차", "오후반차", "병가", "경조휴가", "공가", "특별휴가"]},
                {"name": "start_date", "label": "시작일", "type": "date", "required": True},
                {"name": "end_date", "label": "종료일", "type": "date", "required": True},
                {"name": "days", "label": "사용일수", "type": "number", "required": True},
                {"name": "reason", "label": "사유", "type": "textarea", "required": True},
                {"name": "emergency_contact", "label": "비상연락처", "type": "text", "required": False},
                {"name": "handover", "label": "업무 인수인계", "type": "textarea", "required": False},
                {"name": "approval_status", "label": "승인상태", "type": "select", "required": True, "options": ["신청", "승인", "반려"]},
            ],
        ),
        (
            "주간업무보고", "주간 업무 현황 보고", "📊", "#6366f1",
            [
                {"name": "week_range", "label": "보고 기간", "type": "text", "required": True},
                {"name": "author", "label": "작성자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "completed", "label": "금주 완료 업무", "type": "textarea", "required": True},
                {"name": "in_progress", "label": "진행 중 업무", "type": "textarea", "required": False},
                {"name": "next_week", "label": "차주 계획", "type": "textarea", "required": True},
                {"name": "issues", "label": "이슈/리스크", "type": "textarea", "required": False},
                {"name": "kpi_note", "label": "성과 지표 메모", "type": "textarea", "required": False},
            ],
        ),
        (
            "프로젝트 기획서", "프로젝트 계획 및 제안 양식", "🚀", "#8b5cf6",
            [
                {"name": "project_name", "label": "프로젝트명", "type": "text", "required": True},
                {"name": "pm", "label": "담당자(PM)", "type": "text", "required": True},
                {"name": "start_date", "label": "시작일", "type": "date", "required": True},
                {"name": "end_date", "label": "종료일", "type": "date", "required": True},
                {"name": "background", "label": "추진 배경", "type": "textarea", "required": True},
                {"name": "objective", "label": "목표", "type": "textarea", "required": True},
                {"name": "scope", "label": "범위", "type": "textarea", "required": True},
                {"name": "budget", "label": "예산(원)", "type": "number", "required": False},
                {"name": "milestones", "label": "주요 마일스톤", "type": "textarea", "required": False},
                {"name": "risks", "label": "리스크/대응방안", "type": "textarea", "required": False},
                {"name": "expected_outcome", "label": "기대 효과", "type": "textarea", "required": False},
            ],
        ),
        (
            "인수인계서", "업무 인수인계 양식", "🤝", "#f97316",
            [
                {"name": "handover_date", "label": "인수인계일", "type": "date", "required": True},
                {"name": "from_person", "label": "인계자", "type": "text", "required": True},
                {"name": "to_person", "label": "인수자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "responsibilities", "label": "담당 업무 목록", "type": "textarea", "required": True},
                {"name": "ongoing", "label": "진행 중 업무", "type": "textarea", "required": True},
                {"name": "pending", "label": "미결/보류 사항", "type": "textarea", "required": False},
                {"name": "contacts", "label": "주요 연락처", "type": "textarea", "required": False},
                {"name": "files_location", "label": "관련 자료 위치", "type": "textarea", "required": False},
                {"name": "notes", "label": "특이사항/주의점", "type": "textarea", "required": False},
            ],
        ),
        (
            "교육훈련보고서", "교육/세미나 참가 보고", "🎓", "#a855f7",
            [
                {"name": "title", "label": "교육명", "type": "text", "required": True},
                {"name": "trainee", "label": "참가자", "type": "text", "required": True},
                {"name": "start_date", "label": "시작일", "type": "date", "required": True},
                {"name": "end_date", "label": "종료일", "type": "date", "required": True},
                {"name": "institution", "label": "교육기관/장소", "type": "text", "required": False},
                {"name": "hours", "label": "교육시간", "type": "number", "required": False},
                {"name": "content_summary", "label": "교육 내용 요약", "type": "textarea", "required": True},
                {"name": "key_takeaways", "label": "핵심 시사점", "type": "textarea", "required": True},
                {"name": "application_plan", "label": "업무 적용 계획", "type": "textarea", "required": False},
                {"name": "cost", "label": "교육비(원)", "type": "number", "required": False},
            ],
        ),
        (
            "검수보고서", "납품물/결과물 검수 양식", "🔍", "#ef4444",
            [
                {"name": "title", "label": "검수 건명", "type": "text", "required": True},
                {"name": "inspector", "label": "검수자", "type": "text", "required": True},
                {"name": "inspect_date", "label": "검수일", "type": "date", "required": True},
                {"name": "contractor", "label": "납품업체/담당자", "type": "text", "required": False},
                {"name": "deliverables", "label": "납품 내역", "type": "textarea", "required": True},
                {"name": "criteria", "label": "검수 기준", "type": "textarea", "required": True},
                {"name": "result", "label": "검수 결과", "type": "select", "required": True, "options": ["합격", "조건부합격", "불합격", "재검수"]},
                {"name": "defects", "label": "하자/보완사항", "type": "textarea", "required": False},
                {"name": "opinion", "label": "검수 의견", "type": "textarea", "required": False},
            ],
        ),
        (
            "견적요청서", "견적 요청/비교 양식", "💼", "#64748b",
            [
                {"name": "title", "label": "요청 건명", "type": "text", "required": True},
                {"name": "requester", "label": "요청자", "type": "text", "required": True},
                {"name": "request_date", "label": "요청일", "type": "date", "required": True},
                {"name": "deadline", "label": "회신 기한", "type": "date", "required": False},
                {"name": "vendor", "label": "견적 업체", "type": "text", "required": False},
                {"name": "items", "label": "요청 품목/서비스", "type": "textarea", "required": True},
                {"name": "quantity", "label": "수량/규모", "type": "text", "required": False},
                {"name": "budget_range", "label": "예산 범위", "type": "text", "required": False},
                {"name": "conditions", "label": "요구 조건", "type": "textarea", "required": False},
                {"name": "notes", "label": "비고", "type": "textarea", "required": False},
            ],
        ),
        (
            "시말서", "사고/위반 경위 보고", "⚠️", "#dc2626",
            [
                {"name": "author", "label": "작성자", "type": "text", "required": True},
                {"name": "department", "label": "부서", "type": "text", "required": False},
                {"name": "incident_date", "label": "발생일시", "type": "datetime", "required": True},
                {"name": "incident_type", "label": "사건 유형", "type": "select", "required": True, "options": ["안전사고", "업무과실", "규정위반", "장비손상", "기타"]},
                {"name": "description", "label": "사건 경위", "type": "textarea", "required": True},
                {"name": "cause", "label": "원인 분석", "type": "textarea", "required": True},
                {"name": "damage", "label": "피해/손실 내역", "type": "textarea", "required": False},
                {"name": "corrective_action", "label": "시정 조치", "type": "textarea", "required": True},
                {"name": "prevention", "label": "재발 방지 대책", "type": "textarea", "required": True},
            ],
        ),
    ]
    with get_db() as conn:
        for name, desc, emoji, color, fields in presets:
            existing = conn.execute(
                "SELECT id FROM form_templates WHERE name=? AND profile_id=0", (name,)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color) VALUES (0, ?, ?, ?, ?, ?)",
                    (name, desc, json.dumps(fields, ensure_ascii=False), emoji, color),
                )


# ── Audit log helper ──
def _audit_log(conn, entity_type: str, entity_id: int, action: str, changes: dict = None):
    """Insert a lightweight audit record. JM is single-user so no profile_id filter."""
    conn.execute(
        "INSERT INTO audit_log (entity_type, entity_id, action, changes_json) VALUES (?,?,?,?)",
        (entity_type, entity_id, action, json.dumps(changes or {}, ensure_ascii=False)),
    )


# ── Utility functions ──
from common.constants import PRIORITY_MAP, REPEAT_MAP, WEEKDAY_NAMES, ROLE_COLORS
from common.recurrence import next_occurrence, expand_recurring_events

# ── Google Calendar OAuth2 (constants and helpers from common.gcal) ──


def _gcal_get_cal_id(profile_id: int) -> str:
    """Look up the calendar_id for a profile from the DB."""
    with get_db() as conn:
        row = conn.execute("SELECT calendar_id FROM gcal_tokens WHERE profile_id=?", (profile_id,)).fetchone()
    return row["calendar_id"] if row else "primary"


async def _gcal_refresh_token(profile_id: int) -> Optional[str]:
    """Refresh and return a valid access token for the given profile."""
    with get_db() as conn:
        return await _common_gcal_refresh_token(conn, profile_id)


async def _gcal_fetch_events(profile_id: int, time_min: str, time_max: str) -> list:
    token = await _gcal_refresh_token(profile_id)
    if not token:
        return []
    cal_id = _gcal_get_cal_id(profile_id)
    return await _common_gcal_fetch_events(token, cal_id, time_min, time_max)


async def _gcal_push_event(profile_id: int, title: str, start_time: str, end_time: str = "") -> str:
    """Create an event in Google Calendar. Returns the gcal event ID or empty string."""
    token = await _gcal_refresh_token(profile_id)
    if not token:
        return ""
    cal_id = _gcal_get_cal_id(profile_id)
    return await _common_gcal_push_event(token, cal_id, title, start_time, end_time)


async def _gcal_update_event(profile_id: int, gcal_id: str, title: str, start_time: str, end_time: str = ""):
    """Update an existing event in Google Calendar."""
    if not gcal_id:
        return
    token = await _gcal_refresh_token(profile_id)
    if not token:
        return
    cal_id = _gcal_get_cal_id(profile_id)
    await _common_gcal_update_event(token, cal_id, gcal_id, title, start_time, end_time)


async def _gcal_delete_event(profile_id: int, gcal_id: str):
    """Delete an event from Google Calendar."""
    if not gcal_id:
        return
    token = await _gcal_refresh_token(profile_id)
    if not token:
        return
    cal_id = _gcal_get_cal_id(profile_id)
    await _common_gcal_delete_event(token, cal_id, gcal_id)


from common.stats import get_stats, get_weekly_chart_data, week_number_in_month, get_week_range


def calc_dday(target_date_str: str) -> int:
    try:
        target = datetime.strptime(target_date_str, "%Y-%m-%d").date()
        return (target - date.today()).days
    except (ValueError, TypeError):
        return 0


def get_profile_id(request: Request) -> int:
    return 1


def get_user_setting(conn, profile_id, key: str, default: str = "") -> str:
    row = conn.execute(
        "SELECT value FROM user_settings WHERE profile_id=? AND key=?",
        (str(profile_id), key),
    ).fetchone()
    return row[0] if row else default


def set_user_setting(conn, profile_id, key: str, value: str):
    conn.execute(
        "INSERT OR REPLACE INTO user_settings (profile_id, key, value) VALUES (?, ?, ?)",
        (str(profile_id), key, value),
    )


def redirect(request: Request, url: str):
    if request.headers.get("HX-Request"):
        return HTMLResponse("", headers={"HX-Redirect": url})
    return RedirectResponse(url, status_code=303)


def get_profile_name(request: Request) -> str:
    return "정미"


def _run_automation_rules(conn, pid, today_str):
    rules = conn.execute(
        "SELECT * FROM automation_rules WHERE profile_id=? AND enabled=1", (pid,)
    ).fetchall()
    today_date = date.today()
    for rule in rules:
        r = dict(rule)
        if r['last_run'] == today_str:
            continue
        tc = json.loads(r['trigger_config'])
        should_run = False
        if r['trigger_type'] == 'daily':
            should_run = True
        elif r['trigger_type'] == 'weekly':
            should_run = (today_date.weekday() == tc.get('weekday', 0))
        elif r['trigger_type'] == 'monthly':
            should_run = (today_date.day == tc.get('day', 1))
        if not should_run:
            continue
        ac = json.loads(r['action_config'])
        if r['action_type'] == 'create_todo':
            max_order = conn.execute(
                "SELECT COALESCE(MAX(sort_order),0) FROM todos WHERE profile_id=?", (pid,)
            ).fetchone()[0]
            conn.execute(
                "INSERT INTO todos (profile_id, title, description, priority, category_id, due_date, tags, sort_order) VALUES (?,?,?,?,?,?,?,?)",
                (pid, ac.get('title', ''), ac.get('description', ''),
                 ac.get('priority', 2), ac.get('category_id') or None,
                 today_str, ac.get('tags', ''), max_order + 1))
        conn.execute("UPDATE automation_rules SET last_run=? WHERE id=?", (today_str, r['id']))


# ── Routes: Dashboard ──
@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, plan_view: str = "week", plan_offset: int = 0):
    pid = get_profile_id(request)
    today = date.today()
    today_str = today.isoformat()

    with get_db() as conn:
        _run_automation_rules(conn, pid, today_str)
        week_start, week_end = get_week_range()
        stats = get_stats(conn, pid)

        today_todos = conn.execute("""
            SELECT t.*, c.name as category_name, c.color as category_color
            FROM todos t LEFT JOIN categories c ON t.category_id = c.id
            WHERE ((t.due_date <= ? AND t.completed = 0) OR (t.completed = 1 AND date(t.completed_at) = ?))
              AND t.profile_id = ?
            ORDER BY t.completed ASC, t.priority ASC, t.sort_order ASC
            LIMIT 10
        """, (today_str, today_str, pid)).fetchall()

        week_events = conn.execute("""
            SELECT e.*, c.name as category_name
            FROM events e LEFT JOIN categories c ON e.category_id = c.id
            WHERE date(e.start_time) BETWEEN ? AND ?
              AND e.profile_id = ?
            ORDER BY e.start_time ASC
            LIMIT 5
        """, (week_start.isoformat(), week_end.isoformat(), pid)).fetchall()

        recent_memos = conn.execute("""
            SELECT m.*, c.name as category_name, c.color as category_color
            FROM memos m LEFT JOIN categories c ON m.category_id = c.id
            WHERE m.profile_id = ? ORDER BY m.created_at DESC LIMIT 3
        """, (pid,)).fetchall()

        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()

        recent_notices = conn.execute("""
            SELECT * FROM notices ORDER BY pinned DESC, created_at DESC LIMIT 2
        """).fetchall()

        today_worklogs = conn.execute("""
            SELECT wl.*, c.name as category_name, c.color as category_color
            FROM work_logs wl LEFT JOIN categories c ON wl.category_id = c.id
            WHERE wl.log_date = ?
            ORDER BY wl.created_at DESC
        """, (today_str,)).fetchall()
        today_work_hours = sum((dict(l).get("hours", 0) or 0) for l in today_worklogs)

        project_progress = conn.execute("""
            SELECT c.name, c.color,
                   COUNT(CASE WHEN t.completed=0 THEN 1 END) as pending,
                   COUNT(CASE WHEN t.completed=1 THEN 1 END) as done,
                   COUNT(*) as total
            FROM categories c
            LEFT JOIN todos t ON t.category_id = c.id AND t.profile_id = ?
            GROUP BY c.id HAVING total > 0
            ORDER BY c.sort_order
        """, (pid,)).fetchall()

        monday = today - timedelta(days=today.weekday())
        sunday = monday + timedelta(days=6)
        time_budgets_raw = conn.execute("""
            SELECT w.category_id, c.name, c.color, COALESCE(SUM(w.hours), 0) as used,
                   COALESCE(tb.weekly_hours, 0) as budget
            FROM work_logs w
            LEFT JOIN categories c ON w.category_id = c.id
            LEFT JOIN time_budgets tb ON tb.category_id = w.category_id AND tb.profile_id = ?
            WHERE w.log_date >= ? AND w.log_date <= ?
            GROUP BY w.category_id
        """, (pid, monday.isoformat(), sunday.isoformat())).fetchall()
        time_budgets = [dict(r) for r in time_budgets_raw]
        over_budget = [h for h in time_budgets if h["budget"] > 0 and h["used"] > h["budget"]]

        # Plan view data
        plan_data = {}
        if plan_view == "month":
            m = today.month + plan_offset
            y = today.year
            while m < 1:
                m += 12; y -= 1
            while m > 12:
                m -= 12; y += 1
            _, days_in_month = cal_mod.monthrange(y, m)
            month_start = date(y, m, 1)
            month_end = date(y, m, days_in_month)
            plan_todos = conn.execute("""
                SELECT t.*, c.name as category_name, c.color as category_color
                FROM todos t LEFT JOIN categories c ON t.category_id = c.id
                WHERE t.due_date BETWEEN ? AND ? AND t.profile_id = ?
                ORDER BY t.due_date ASC, t.priority ASC, t.sort_order ASC
            """, (month_start.isoformat(), month_end.isoformat(), pid)).fetchall()
            start_weekday = month_start.weekday()
            cal_weeks = []
            current_week = [None] * start_weekday
            for day_num in range(1, days_in_month + 1):
                current_week.append(date(y, m, day_num))
                if len(current_week) == 7:
                    cal_weeks.append(current_week); current_week = []
            if current_week:
                current_week += [None] * (7 - len(current_week))
                cal_weeks.append(current_week)
            todos_by_date: dict = {}
            for t in plan_todos:
                td = dict(t)
                dd = td.get("due_date", "")
                if dd:
                    todos_by_date.setdefault(dd, []).append(td)
            plan_data = {
                "cal_weeks": cal_weeks, "todos_by_date": todos_by_date,
                "nav_label": f"{y}년 {m}월", "reset_label": "이번달",
                "total_count": len(plan_todos),
                "done_count": sum(1 for t in plan_todos if t["completed"]),
                "weekday_names": WEEKDAY_NAMES,
            }
        else:
            plan_monday = today - timedelta(days=today.weekday()) + timedelta(weeks=plan_offset)
            plan_sunday = plan_monday + timedelta(days=6)
            week_num = week_number_in_month(plan_monday)
            plan_todos = conn.execute("""
                SELECT t.*, c.name as category_name, c.color as category_color
                FROM todos t LEFT JOIN categories c ON t.category_id = c.id
                WHERE t.due_date BETWEEN ? AND ? AND t.profile_id = ?
                ORDER BY t.due_date ASC, t.priority ASC, t.sort_order ASC
            """, (plan_monday.isoformat(), plan_sunday.isoformat(), pid)).fetchall()
            week_days = []
            for i in range(7):
                d = plan_monday + timedelta(days=i)
                day_todos = [dict(t) for t in plan_todos if t["due_date"] == d.isoformat()]
                week_days.append({
                    "date": d, "date_str": d.isoformat(),
                    "label": f"{WEEKDAY_NAMES[i]} {d.strftime('%m/%d')}",
                    "short_label": WEEKDAY_NAMES[i],
                    "is_today": d == today, "is_weekend": i >= 5,
                    "todos": day_todos,
                })
            plan_data = {
                "week_days": week_days,
                "nav_label": f"{plan_monday.year}년 {plan_monday.month}월 {week_num}주차 ({plan_monday.strftime('%m.%d')} ~ {plan_sunday.strftime('%m.%d')})",
                "reset_label": "오늘",
                "total_count": len(plan_todos),
                "done_count": sum(1 for t in plan_todos if t["completed"]),
            }

    return render(request, "dashboard.html", {
        "page": "dashboard",
        "stats": stats,
        "today_todos": [dict(r) for r in today_todos],
        "week_events": [dict(r) for r in week_events],
        "recent_memos": [dict(r) for r in recent_memos],
        "recent_notices": [dict(r) for r in recent_notices],
        "today_worklogs": [dict(r) for r in today_worklogs],
        "today_work_hours": today_work_hours,
        "categories": [dict(r) for r in categories],
        "project_progress": [dict(r) for r in project_progress],
        "time_budgets": time_budgets,
        "over_budget": over_budget,
        "priority_map": PRIORITY_MAP,
        "plan_view": plan_view,
        "plan_offset": plan_offset,
        **plan_data,
    })


# ── Routes: Todos ──
@app.get("/todos", response_class=HTMLResponse)
async def todos_page(request: Request, filter: str = "all", category_id: Optional[int] = None, assignee: Optional[str] = None, energy: Optional[int] = None):
    pid = get_profile_id(request)
    with get_db() as conn:
        today_str = date.today().isoformat()
        where = "1=1"
        params: list = []

        if filter == "completed":
            where = "t.completed = 1"
        elif filter == "active":
            where = "t.completed = 0"
        elif filter == "overdue":
            where = "t.completed = 0 AND t.due_date < ? AND t.due_date IS NOT NULL"
            params.append(today_str)
        else:
            where = "1=1"

        where += " AND t.profile_id = ?"
        params.append(pid)

        if category_id:
            where += " AND t.category_id = ?"
            params.append(category_id)

        if assignee:
            where += " AND t.assignee = ?"
            params.append(assignee)

        if energy in (1, 2, 3):
            where += " AND t.energy_level = ?"
            params.append(energy)

        profiles = conn.execute("SELECT * FROM work_profiles ORDER BY id").fetchall()

        todos = conn.execute(f"""
            SELECT t.*, c.name as category_name, c.color as category_color
            FROM todos t LEFT JOIN categories c ON t.category_id = c.id
            WHERE {where}
            ORDER BY t.completed ASC,
                     CASE WHEN t.due_date IS NULL THEN 1 ELSE 0 END,
                     t.due_date ASC, t.priority ASC, t.sort_order ASC
        """, params).fetchall()

        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()

        from collections import OrderedDict
        grouped: OrderedDict[str, list] = OrderedDict()
        for t in todos:
            td = dict(t)
            subs = conn.execute(
                "SELECT * FROM subtasks WHERE todo_id=? ORDER BY sort_order, id", (td["id"],)
            ).fetchall()
            td["subtasks"] = [dict(s) for s in subs]
            key = td.get("due_date") or ""
            grouped.setdefault(key, []).append(td)

    return render(request, "todos.html", {
        "page": "todos",
        "todo_groups": grouped,
        "todo_count": sum(len(v) for v in grouped.values()),
        "categories": [dict(c) for c in categories],
        "profiles": [dict(p) for p in profiles],
        "current_filter": filter,
        "current_category_id": category_id,
        "current_assignee": assignee,
        "current_energy": energy,
        "priority_map": PRIORITY_MAP,
        "repeat_map": REPEAT_MAP,
    })


@app.post("/todos", response_class=HTMLResponse)
async def create_todo(request: Request,
                      title: str = Form(...),
                      description: str = Form(""),
                      due_date: str = Form(""),
                      priority: int = Form(2),
                      category_id: str = Form(""),
                      tags: str = Form(""),
                      repeat_type: str = Form("none"),
                      recurrence_end: str = Form(""),
                      assignee: str = Form(""),
                      energy_level: int = Form(2)):
    title = clamp_text(fix_mojibake(title), 200).strip()
    if not title:
        return redirect(request, "/todos")
    description = clamp_text(fix_mojibake(description), 2000)
    assignee = clamp_text(fix_mojibake(assignee), 100)
    priority = clamp_priority(priority)
    due_date = validate_date_str(due_date)
    recurrence_end = validate_date_str(recurrence_end) or ""
    energy_level = max(1, min(3, energy_level))

    tag_list = [t.strip() for t in fix_mojibake(tags).split(",") if t.strip()] if tags else []
    cat_id = int(category_id) if category_id else None

    pid = get_profile_id(request)
    with get_db() as conn:
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM todos").fetchone()[0]
        cur = conn.execute("""
            INSERT INTO todos (title, description, due_date, priority, category_id, tags, repeat_type, recurrence_end, assignee, sort_order, profile_id, energy_level)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (title, description, due_date, priority, cat_id, json.dumps(tag_list), repeat_type, recurrence_end, assignee, max_order + 1, pid, energy_level))
        _audit_log(conn, "todo", cur.lastrowid, "create", {"title": title})

    return redirect(request, "/todos")


@app.post("/todos/{todo_id}/toggle", response_class=HTMLResponse)
async def toggle_todo(request: Request, todo_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        todo = conn.execute("SELECT * FROM todos WHERE id=? AND profile_id=?", (todo_id, pid)).fetchone()
        if not todo:
            raise HTTPException(404)
        new_status = 0 if todo["completed"] else 1
        completed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if new_status else None
        conn.execute(
            "UPDATE todos SET completed=?, completed_at=?, updated_at=datetime('now','localtime') WHERE id=?",
            (new_status, completed_at, todo_id),
        )
        _audit_log(conn, "todo", todo_id, "complete" if new_status else "uncomplete", {"title": todo["title"]})

        # Handle repeat
        if new_status == 1 and todo["repeat_type"] != "none" and todo["due_date"]:
            nxt = next_occurrence(todo["due_date"], todo["repeat_type"])
            rec_end = todo["recurrence_end"] if "recurrence_end" in todo.keys() else ""
            if nxt and (not rec_end or nxt <= rec_end):
                max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM todos").fetchone()[0]
                conn.execute("""
                    INSERT INTO todos (title, description, due_date, priority, category_id, tags, repeat_type, recurrence_end, assignee, sort_order, profile_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (todo["title"], todo["description"], nxt,
                      todo["priority"], todo["category_id"], todo["tags"],
                      todo["repeat_type"], rec_end, todo["assignee"], max_order + 1, todo["profile_id"]))

    if request.headers.get("HX-Request"):
        with get_db() as conn:
            updated = conn.execute(
                "SELECT t.*, c.name as category_name, c.color as category_color FROM todos t LEFT JOIN categories c ON t.category_id=c.id WHERE t.id=? AND t.profile_id=?",
                (todo_id, pid),
            ).fetchone()
            if updated:
                td = dict(updated)
                td["subtasks"] = [dict(s) for s in conn.execute("SELECT * FROM subtasks WHERE todo_id=? ORDER BY sort_order, id", (todo_id,)).fetchall()]
                return render(request, "partials/todo_item.html", {"todo": td, "priority_map": PRIORITY_MAP, "repeat_map": REPEAT_MAP, "today": date.today()})
        return HTMLResponse("")

    referer = request.headers.get("HX-Current-URL") or request.headers.get("referer") or ""
    if "/calendar" in referer:
        return redirect(request, "/calendar")
    if "plan_view" in referer or referer.rstrip("/").endswith(":8002") or referer.endswith("/?"):
        return redirect(request, "/")
    return redirect(request, "/todos")


@app.get("/todos/{todo_id}/edit", response_class=HTMLResponse)
async def edit_todo_form(request: Request, todo_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        todo = conn.execute("SELECT * FROM todos WHERE id=? AND profile_id=?", (todo_id, pid)).fetchone()
        if not todo:
            raise HTTPException(404)
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    return render(request, "partials/todo_edit_form.html", {
        "todo": dict(todo),
        "categories": [dict(c) for c in categories],
        "priority_map": PRIORITY_MAP,
        "repeat_map": REPEAT_MAP,
    })


@app.put("/todos/{todo_id}", response_class=HTMLResponse)
async def update_todo(request: Request, todo_id: int,
                      title: str = Form(...),
                      description: str = Form(""),
                      due_date: str = Form(""),
                      priority: int = Form(2),
                      category_id: str = Form(""),
                      tags: str = Form(""),
                      repeat_type: str = Form("none"),
                      recurrence_end: str = Form(""),
                      assignee: str = Form(""),
                      energy_level: int = Form(2)):
    title = clamp_text(fix_mojibake(title), 200)
    description = clamp_text(fix_mojibake(description), 2000)
    assignee = clamp_text(fix_mojibake(assignee), 100)
    priority = clamp_priority(priority)
    due_date = validate_date_str(due_date)
    recurrence_end = validate_date_str(recurrence_end) or ""
    energy_level = max(1, min(3, energy_level))
    tag_list = [t.strip() for t in fix_mojibake(tags).split(",") if t.strip()] if tags else []
    cat_id = int(category_id) if category_id else None

    pid = get_profile_id(request)
    with get_db() as conn:
        old = conn.execute("SELECT title, description, due_date, priority FROM todos WHERE id=? AND profile_id=?", (todo_id, pid)).fetchone()
        conn.execute("""
            UPDATE todos SET title=?, description=?, due_date=?, priority=?, category_id=?,
                   tags=?, repeat_type=?, recurrence_end=?, assignee=?, energy_level=?, updated_at=datetime('now','localtime')
            WHERE id=? AND profile_id=?
        """, (title, description, due_date, priority, cat_id, json.dumps(tag_list), repeat_type, recurrence_end, assignee, energy_level, todo_id, pid))
        changes = {}
        if old:
            if old["title"] != title:
                changes["title"] = {"old": old["title"], "new": title}
            if old["description"] != description:
                changes["description"] = {"old": old["description"][:50], "new": description[:50]}
            if old["due_date"] != due_date:
                changes["due_date"] = {"old": old["due_date"], "new": due_date}
            if old["priority"] != priority:
                changes["priority"] = {"old": old["priority"], "new": priority}
        _audit_log(conn, "todo", todo_id, "update", changes)

    return redirect(request, "/todos")


@app.delete("/todos/{todo_id}", response_class=HTMLResponse)
async def delete_todo(request: Request, todo_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        old = conn.execute("SELECT title FROM todos WHERE id=? AND profile_id=?", (todo_id, pid)).fetchone()
        conn.execute("DELETE FROM todos WHERE id=? AND profile_id=?", (todo_id, pid))
        _audit_log(conn, "todo", todo_id, "delete", {"title": old["title"]} if old else {})
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/todos")


@app.post("/todos/reorder", response_class=HTMLResponse)
async def reorder_todos(request: Request):
    pid = get_profile_id(request)
    body = await request.json()
    order = body.get("order", [])
    with get_db() as conn:
        for idx, tid in enumerate(order):
            conn.execute("UPDATE todos SET sort_order=? WHERE id=? AND profile_id=?", (idx, int(tid), pid))
    return JSONResponse({"ok": True})


@app.post("/todos/bulk", response_class=HTMLResponse)
async def bulk_todo_action(request: Request):
    pid = get_profile_id(request)
    body = await request.json()
    action = body.get("action")
    ids = [int(i) for i in body.get("ids", []) if str(i).isdigit()]
    if not ids or action not in ("complete", "delete"):
        return JSONResponse({"ok": False})
    placeholders = ",".join("?" * len(ids))
    with get_db() as conn:
        if action == "complete":
            conn.execute(f"UPDATE todos SET completed=1 WHERE id IN ({placeholders}) AND profile_id=?", (*ids, pid))
        elif action == "delete":
            conn.execute(f"DELETE FROM subtasks WHERE todo_id IN ({placeholders})", ids)
            conn.execute(f"DELETE FROM todos WHERE id IN ({placeholders}) AND profile_id=?", (*ids, pid))
    return JSONResponse({"ok": True})


# ── Subtasks ──
@app.post("/todos/{todo_id}/subtasks", response_class=HTMLResponse)
async def add_subtask(request: Request, todo_id: int, title: str = Form(...)):
    title = clamp_text(fix_mojibake(title), 200)
    with get_db() as conn:
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM subtasks WHERE todo_id=?", (todo_id,)).fetchone()[0]
        conn.execute("INSERT INTO subtasks (todo_id, title, sort_order) VALUES (?,?,?)", (todo_id, title, max_order + 1))
    return redirect(request, "/todos")


@app.post("/subtasks/{sub_id}/toggle", response_class=HTMLResponse)
async def toggle_subtask(request: Request, sub_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        sub = conn.execute("SELECT * FROM subtasks WHERE id=?", (sub_id,)).fetchone()
        if sub:
            conn.execute("UPDATE subtasks SET completed=? WHERE id=?", (0 if sub["completed"] else 1, sub_id))
            if request.headers.get("HX-Request"):
                todo_id = sub["todo_id"]
                updated = conn.execute(
                    "SELECT t.*, c.name as category_name, c.color as category_color FROM todos t LEFT JOIN categories c ON t.category_id=c.id WHERE t.id=? AND t.profile_id=?",
                    (todo_id, pid),
                ).fetchone()
                if updated:
                    td = dict(updated)
                    td["subtasks"] = [dict(s) for s in conn.execute("SELECT * FROM subtasks WHERE todo_id=? ORDER BY sort_order, id", (todo_id,)).fetchall()]
                    return render(request, "partials/todo_item.html", {"todo": td, "priority_map": PRIORITY_MAP, "repeat_map": REPEAT_MAP, "today": date.today()})
    return redirect(request, "/todos")


@app.delete("/subtasks/{sub_id}", response_class=HTMLResponse)
async def delete_subtask(request: Request, sub_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM subtasks WHERE id=?", (sub_id,))
    return HTMLResponse("")


# ── Routes: Todo Templates ──
@app.get("/todo-templates", response_class=HTMLResponse)
async def todo_templates_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpls = [dict(r) for r in conn.execute(
            "SELECT * FROM todo_templates WHERE profile_id=? ORDER BY created_at DESC", (pid,)
        ).fetchall()]
        categories = [dict(r) for r in conn.execute(
            "SELECT * FROM categories ORDER BY sort_order"
        ).fetchall()]
    for t in tpls:
        t['items'] = json.loads(t['items_json'])
    return render(request, "todo_templates.html", {
        "page": "todo-templates", "templates": tpls, "categories": categories
    })


@app.post("/todo-templates", response_class=HTMLResponse)
async def create_todo_template(request: Request):
    pid = get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    items_json = form.get("items_json", "[]")
    if not name:
        return redirect(request, "/todo-templates")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO todo_templates (profile_id, name, items_json) VALUES (?,?,?)",
            (pid, name, items_json))
    return redirect(request, "/todo-templates")


@app.post("/todo-templates/{tpl_id}/apply", response_class=HTMLResponse)
async def apply_todo_template(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute("SELECT * FROM todo_templates WHERE id=? AND profile_id=?", (tpl_id, pid)).fetchone()
        if not tpl:
            return redirect(request, "/todo-templates")
        items = json.loads(tpl['items_json'])
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM todos WHERE profile_id=?", (pid,)).fetchone()[0]
        today = date.today().isoformat()
        for i, item in enumerate(items):
            conn.execute(
                "INSERT INTO todos (profile_id, title, description, priority, category_id, due_date, tags, sort_order) VALUES (?,?,?,?,?,?,?,?)",
                (pid, item.get('title', ''), item.get('description', ''), item.get('priority', 2),
                 item.get('category_id') or None, today, item.get('tags', ''), max_order + i + 1))
    return redirect(request, "/todos")


@app.delete("/todo-templates/{tpl_id}", response_class=HTMLResponse)
async def delete_todo_template(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM todo_templates WHERE id=? AND profile_id=?", (tpl_id, pid))
    return redirect(request, "/todo-templates")


@app.post("/todo-templates/from-todos", response_class=HTMLResponse)
async def create_template_from_todos(request: Request):
    pid = get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    todo_ids = form.getlist("todo_ids")
    if not name or not todo_ids:
        return redirect(request, "/todos")
    placeholders = ",".join("?" * len(todo_ids))
    with get_db() as conn:
        todos = conn.execute(
            f"SELECT title, description, priority, category_id, tags FROM todos WHERE id IN ({placeholders}) AND profile_id=?",
            (*[int(i) for i in todo_ids], pid)).fetchall()
        items = [{"title": t["title"], "description": t["description"] or "", "priority": t["priority"],
                  "category_id": t["category_id"], "tags": t["tags"] or ""} for t in todos]
        conn.execute("INSERT INTO todo_templates (profile_id, name, items_json) VALUES (?,?,?)",
                     (pid, name, json.dumps(items, ensure_ascii=False)))
    return redirect(request, "/todo-templates")


# ── Routes: Automation Rules ──
@app.get("/automations", response_class=HTMLResponse)
async def automations_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        rules = [dict(r) for r in conn.execute(
            "SELECT * FROM automation_rules WHERE profile_id=? ORDER BY created_at DESC", (pid,)
        ).fetchall()]
        categories = [dict(r) for r in conn.execute(
            "SELECT * FROM categories ORDER BY sort_order"
        ).fetchall()]
    for ru in rules:
        ru['_tc'] = json.loads(ru.get('trigger_config') or '{}')
        ru['_ac'] = json.loads(ru.get('action_config') or '{}')
    return render(request, "automations.html", {
        "page": "automations", "rules": rules, "categories": categories,
    })


@app.post("/automations", response_class=HTMLResponse)
async def create_automation(request: Request):
    pid = get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    if not name:
        return redirect(request, "/automations")
    trigger_type = form.get("trigger_type", "weekly")
    trigger_config = form.get("trigger_config", "{}")
    action_config = form.get("action_config", "{}")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO automation_rules (profile_id, name, trigger_type, trigger_config, action_type, action_config) VALUES (?,?,?,?,?,?)",
            (pid, name, trigger_type, trigger_config, "create_todo", action_config))
    return redirect(request, "/automations")


@app.post("/automations/{rule_id}/toggle", response_class=HTMLResponse)
async def toggle_automation(request: Request, rule_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute(
            "UPDATE automation_rules SET enabled = CASE WHEN enabled=1 THEN 0 ELSE 1 END WHERE id=? AND profile_id=?",
            (rule_id, pid))
    return redirect(request, "/automations")


@app.delete("/automations/{rule_id}", response_class=HTMLResponse)
async def delete_automation(request: Request, rule_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM automation_rules WHERE id=? AND profile_id=?", (rule_id, pid))
    return redirect(request, "/automations")


# ── Korean Holidays (imported from common.holidays) ──


# ── Routes: Calendar ──
@app.get("/calendar", response_class=HTMLResponse)
async def calendar_page(request: Request, year: Optional[int] = None, month: Optional[int] = None):
    today = date.today()
    y = year or today.year
    m = month or today.month

    if m < 1:
        m = 12
        y -= 1
    elif m > 12:
        m = 1
        y += 1

    first_day = date(y, m, 1)
    _, days_in_month = cal_mod.monthrange(y, m)
    start_weekday = first_day.weekday()  # 0=Mon

    month_start = first_day.isoformat()
    month_end = date(y, m, days_in_month).isoformat()

    pid = get_profile_id(request)
    with get_db() as conn:
        # Fetch non-recurring events in range + all recurring events that started on or before month end
        events = conn.execute("""
            SELECT e.*, c.name as category_name
            FROM events e LEFT JOIN categories c ON e.category_id = c.id
            WHERE e.profile_id = ?
              AND (
                (COALESCE(e.recurrence, '') = '' AND date(e.start_time) BETWEEN ? AND ?)
                OR (COALESCE(e.recurrence, '') != '' AND date(e.start_time) <= ?)
              )
            ORDER BY e.start_time ASC
        """, (pid, month_start, month_end, month_end)).fetchall()

        todos = conn.execute("""
            SELECT t.*, c.name as category_name, c.color as category_color
            FROM todos t LEFT JOIN categories c ON t.category_id = c.id
            WHERE t.due_date BETWEEN ? AND ?
              AND t.profile_id = ?
            ORDER BY t.priority ASC, t.sort_order ASC
        """, (month_start, month_end, pid)).fetchall()

        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()

    # Expand recurring events into individual occurrences
    all_events = expand_recurring_events([dict(ev) for ev in events], month_start, month_end)

    # Group events by date
    events_by_date: dict = {}
    for d in all_events:
        try:
            day_key = d["start_time"][:10]
        except (TypeError, IndexError):
            continue
        events_by_date.setdefault(day_key, []).append(d)

    # Merge Google Calendar events
    gcal_events = await _gcal_fetch_events(pid, month_start, month_end)
    for gev in gcal_events:
        day_key = gev["start_time"][:10] if gev["start_time"] else ""
        if day_key:
            events_by_date.setdefault(day_key, []).append(gev)

    # Group todos by date
    todos_by_date: dict = {}
    for td in todos:
        d = dict(td)
        dd = d.get("due_date", "")
        if dd:
            todos_by_date.setdefault(dd, []).append(d)

    prev_m = m - 1
    prev_y = y
    if prev_m < 1:
        prev_m = 12
        prev_y -= 1
    next_m = m + 1
    next_y = y
    if next_m > 12:
        next_m = 1
        next_y += 1

    return render(request, "calendar.html", {
        "page": "calendar",
        "year": y,
        "month": m,
        "days_in_month": days_in_month,
        "start_weekday": start_weekday,
        "events_by_date": events_by_date,
        "todos_by_date": todos_by_date,
        "categories": [dict(c) for c in categories],
        "prev_year": prev_y,
        "prev_month": prev_m,
        "next_year": next_y,
        "next_month": next_m,
        "today_str": today.isoformat(),
        "holidays_by_date": get_holidays_for_month(y, m),
    })


@app.post("/events", response_class=HTMLResponse)
async def create_event(request: Request,
                       title: str = Form(...),
                       start_time: str = Form(...),
                       end_time: str = Form(""),
                       color: str = Form("#6366f1"),
                       category_id: str = Form(""),
                       memo: str = Form(""),
                       recurrence: str = Form(""),
                       recurrence_end: str = Form(""),
                       also_todo: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    memo = clamp_text(fix_mojibake(memo), 2000)
    start_time = validate_datetime_str(start_time) or datetime.now().strftime("%Y-%m-%dT%H:%M")
    end_time = validate_datetime_str(end_time)
    cat_id = int(category_id) if category_id else None
    recurrence_end = validate_date_str(recurrence_end) or ""

    pid = get_profile_id(request)
    gcal_id = await _gcal_push_event(pid, title, start_time, end_time or "")

    with get_db() as conn:
        conn.execute("""
            INSERT INTO events (title, start_time, end_time, color, category_id, memo, profile_id, gcal_event_id, recurrence, recurrence_end)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (title, start_time, end_time, color, cat_id, memo, pid, gcal_id, recurrence, recurrence_end))

        if also_todo:
            event_date = start_time[:10] if start_time else ""
            due = validate_date_str(event_date)
            max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM todos").fetchone()[0]
            conn.execute("""
                INSERT INTO todos (title, due_date, priority, category_id, repeat_type, sort_order, profile_id)
                VALUES (?, ?, 2, ?, 'none', ?, ?)
            """, (title, due, cat_id, max_order + 1, pid))

    return redirect(request, "/calendar")


@app.get("/events/{event_id}/edit", response_class=HTMLResponse)
async def edit_event_form(request: Request, event_id: int):
    with get_db() as conn:
        event = conn.execute("SELECT * FROM events WHERE id=?", (event_id,)).fetchone()
        if not event:
            raise HTTPException(404)
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    return render(request, "partials/event_edit_form.html", {
        "event": dict(event),
        "categories": [dict(c) for c in categories],
    })


@app.put("/events/{event_id}", response_class=HTMLResponse)
async def update_event(request: Request, event_id: int,
                       title: str = Form(...),
                       start_time: str = Form(...),
                       end_time: str = Form(""),
                       color: str = Form("#6366f1"),
                       category_id: str = Form(""),
                       memo: str = Form(""),
                       recurrence: str = Form(""),
                       recurrence_end: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    memo = clamp_text(fix_mojibake(memo), 2000)
    start_time = validate_datetime_str(start_time) or datetime.now().strftime("%Y-%m-%dT%H:%M")
    end_time = validate_datetime_str(end_time)
    cat_id = int(category_id) if category_id else None
    recurrence_end = validate_date_str(recurrence_end) or ""

    pid = get_profile_id(request)
    with get_db() as conn:
        row = conn.execute("SELECT gcal_event_id FROM events WHERE id=? AND profile_id=?", (event_id, pid)).fetchone()
        gcal_id = row["gcal_event_id"] if row and row["gcal_event_id"] else ""
        conn.execute("""
            UPDATE events SET title=?, start_time=?, end_time=?, color=?, category_id=?, memo=?,
                   recurrence=?, recurrence_end=?, updated_at=datetime('now','localtime')
            WHERE id=? AND profile_id=?
        """, (title, start_time, end_time, color, cat_id, memo, recurrence, recurrence_end, event_id, pid))

    await _gcal_update_event(pid, gcal_id, title, start_time, end_time or "")

    return redirect(request, "/calendar")


@app.delete("/events/{event_id}", response_class=HTMLResponse)
async def delete_event(request: Request, event_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        row = conn.execute("SELECT gcal_event_id FROM events WHERE id=? AND profile_id=?", (event_id, pid)).fetchone()
        gcal_id = row["gcal_event_id"] if row and row["gcal_event_id"] else ""
        conn.execute("DELETE FROM events WHERE id=? AND profile_id=?", (event_id, pid))
    await _gcal_delete_event(pid, gcal_id)
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/calendar")


# ── Routes: Google Calendar Event Edit/Delete ──

@app.get("/events/gcal/{gcal_id:path}/edit", response_class=HTMLResponse)
async def edit_gcal_event_form(request: Request, gcal_id: str):
    """Fetch a Google Calendar event and return an edit form."""
    pid = get_profile_id(request)
    import httpx
    token = await _gcal_refresh_token(pid)
    if not token:
        raise HTTPException(400, "Google Calendar not connected")
    with get_db() as conn:
        row = conn.execute("SELECT calendar_id FROM gcal_tokens WHERE profile_id=?", (pid,)).fetchone()
    cal_id = row["calendar_id"] if row else "primary"
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"{GCAL_API_BASE}/calendars/{cal_id}/events/{gcal_id}",
            headers={"Authorization": f"Bearer {token}"},
        )
    if resp.status_code != 200:
        raise HTTPException(404, "Event not found")
    ev = resp.json()
    # Parse start/end
    start = ev.get("start", {}).get("dateTime", ev.get("start", {}).get("date", ""))
    end = ev.get("end", {}).get("dateTime", ev.get("end", {}).get("date", ""))
    # Trim timezone offset for datetime-local input
    if "T" in start and "+" in start:
        start = start.rsplit("+", 1)[0]
    if "T" in end and "+" in end:
        end = end.rsplit("+", 1)[0]
    event_data = {
        "gcal_id": gcal_id,
        "title": ev.get("summary", ""),
        "start_time": start,
        "end_time": end,
        "memo": ev.get("description", ""),
        "color": "#4285F4",
    }
    return render(request, "partials/gcal_event_edit_form.html", {"event": event_data})


@app.put("/events/gcal/{gcal_id:path}", response_class=HTMLResponse)
async def update_gcal_event(request: Request, gcal_id: str,
                            title: str = Form(...),
                            start_time: str = Form(...),
                            end_time: str = Form(""),
                            memo: str = Form("")):
    pid = get_profile_id(request)
    title = clamp_text(fix_mojibake(title), 200)
    memo = clamp_text(fix_mojibake(memo), 2000)
    import httpx
    token = await _gcal_refresh_token(pid)
    if not token:
        raise HTTPException(400, "Google Calendar not connected")
    with get_db() as conn:
        row = conn.execute("SELECT calendar_id FROM gcal_tokens WHERE profile_id=?", (pid,)).fetchone()
    cal_id = row["calendar_id"] if row else "primary"
    body = {"summary": title, "description": memo}
    if "T" in start_time:
        body["start"] = {"dateTime": start_time + ":00+09:00" if len(start_time) == 16 else start_time, "timeZone": "Asia/Seoul"}
        if end_time and "T" in end_time:
            body["end"] = {"dateTime": end_time + ":00+09:00" if len(end_time) == 16 else end_time, "timeZone": "Asia/Seoul"}
        else:
            body["end"] = body["start"]
    else:
        body["start"] = {"date": start_time[:10]}
        body["end"] = {"date": (end_time[:10] if end_time else start_time[:10])}
    async with httpx.AsyncClient() as client:
        await client.patch(
            f"{GCAL_API_BASE}/calendars/{cal_id}/events/{gcal_id}",
            json=body,
            headers={"Authorization": f"Bearer {token}"},
        )
    return redirect(request, "/calendar")


@app.delete("/events/gcal/{gcal_id:path}", response_class=HTMLResponse)
async def delete_gcal_event(request: Request, gcal_id: str):
    pid = get_profile_id(request)
    import httpx
    token = await _gcal_refresh_token(pid)
    if not token:
        raise HTTPException(400, "Google Calendar not connected")
    with get_db() as conn:
        row = conn.execute("SELECT calendar_id FROM gcal_tokens WHERE profile_id=?", (pid,)).fetchone()
    cal_id = row["calendar_id"] if row else "primary"
    async with httpx.AsyncClient() as client:
        await client.delete(
            f"{GCAL_API_BASE}/calendars/{cal_id}/events/{gcal_id}",
            headers={"Authorization": f"Bearer {token}"},
        )
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/calendar")


# ── Routes: Memos ──
@app.get("/memos", response_class=HTMLResponse)
async def memos_page(request: Request, category_id: Optional[int] = None):
    pid = get_profile_id(request)
    with get_db() as conn:
        params: list = [pid]
        where_extra = ""
        if category_id is not None:
            where_extra = " AND m.category_id = ?"
            params.append(category_id)
        memos = conn.execute(f"""
            SELECT m.*, c.name as category_name, c.color as category_color
            FROM memos m LEFT JOIN categories c ON m.category_id = c.id
            WHERE m.profile_id = ?{where_extra}
            ORDER BY m.created_at DESC
        """, params).fetchall()
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    return render(request, "memos.html", {
        "page": "memos",
        "memos": [dict(m) for m in memos],
        "categories": [dict(c) for c in categories],
        "current_category_id": category_id,
    })


@app.post("/memos", response_class=HTMLResponse)
async def create_memo(request: Request,
                      content: str = Form(...),
                      title: str = Form(""),
                      category_id: str = Form("")):
    content = clamp_text(fix_mojibake(content), 5000)
    title = clamp_text(fix_mojibake(title), 200)
    author = get_profile_name(request)
    pid = get_profile_id(request)
    cat_id = int(category_id) if category_id else None
    with get_db() as conn:
        conn.execute(
            "INSERT INTO memos (author, content, title, category_id, profile_id) VALUES (?, ?, ?, ?, ?)",
            (author, content, title, cat_id, pid),
        )
    return redirect(request, "/memos")


@app.get("/memos/{memo_id}/edit", response_class=HTMLResponse)
async def edit_memo_form(request: Request, memo_id: int):
    with get_db() as conn:
        memo = conn.execute(
            "SELECT m.*, c.name as category_name, c.color as category_color "
            "FROM memos m LEFT JOIN categories c ON m.category_id = c.id WHERE m.id=?",
            (memo_id,),
        ).fetchone()
        if not memo:
            raise HTTPException(404)
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    return templates.TemplateResponse(request, "partials/memo_edit_form.html", {
        "memo": dict(memo), "categories": [dict(c) for c in categories],
    })


@app.put("/memos/{memo_id}", response_class=HTMLResponse)
async def update_memo(request: Request, memo_id: int,
                      title: str = Form(""), content: str = Form(...),
                      category_id: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    content = clamp_text(fix_mojibake(content), 5000)
    cat_id = int(category_id) if category_id else None
    with get_db() as conn:
        conn.execute(
            "UPDATE memos SET title=?, content=?, category_id=? WHERE id=?",
            (title, content, cat_id, memo_id),
        )
        if request.headers.get("HX-Request"):
            memo = conn.execute(
                "SELECT m.*, c.name as category_name, c.color as category_color "
                "FROM memos m LEFT JOIN categories c ON m.category_id = c.id WHERE m.id=?",
                (memo_id,),
            ).fetchone()
            if memo:
                return templates.TemplateResponse(request, "partials/memo_card.html", {
                    "memo": dict(memo),
                })
    return redirect(request, "/memos")


@app.delete("/memos/{memo_id}", response_class=HTMLResponse)
async def delete_memo(request: Request, memo_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM memos WHERE id=? AND profile_id=?", (memo_id, pid))
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/memos")


# ── Routes: Settings (categories + theme) ──
@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
        gcal_row = conn.execute("SELECT * FROM gcal_tokens WHERE profile_id=?", (pid,)).fetchone()
        bg_type = get_user_setting(conn, pid, "bg_type", "none")
        bg_preset = get_user_setting(conn, pid, "bg_preset", "")
        bg_opacity = get_user_setting(conn, pid, "bg_opacity", "0.7")
        budgets = conn.execute(
            "SELECT category_id, weekly_hours FROM time_budgets WHERE profile_id=?", (pid,)
        ).fetchall()
    gcal_connected = gcal_row is not None
    gcal_calendar_id = gcal_row["calendar_id"] if gcal_row else "primary"
    gcal_configured = bool(GCAL_CLIENT_ID)
    budget_map = {b["category_id"]: b["weekly_hours"] for b in budgets}
    time_budget_cats = [dict(c) | {"budget_hours": budget_map.get(c["id"], 0)} for c in categories]
    return render(request, "settings.html", {
        "page": "settings",
        "categories": [dict(c) for c in categories],
        "time_budget_cats": time_budget_cats,
        "gcal_connected": gcal_connected,
        "gcal_calendar_id": gcal_calendar_id,
        "gcal_configured": gcal_configured,
        "current_bg_type": bg_type,
        "current_bg_preset": bg_preset,
        "current_bg_opacity": float(bg_opacity),
    })


@app.post("/settings/categories", response_class=HTMLResponse)
async def create_category(request: Request,
                          name: str = Form(...),
                          color: str = Form("#6366f1")):
    name = clamp_text(fix_mojibake(name), 50)
    with get_db() as conn:
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM categories").fetchone()[0]
        try:
            conn.execute(
                "INSERT INTO categories (name, color, sort_order) VALUES (?, ?, ?)",
                (name, color, max_order + 1),
            )
        except sqlite3.IntegrityError:
            pass  # duplicate name
    return redirect(request, "/settings")


@app.delete("/settings/categories/{cat_id}", response_class=HTMLResponse)
async def delete_category(request: Request, cat_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM categories WHERE id=?", (cat_id,))
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/settings")


@app.get("/settings/time-budgets", response_class=HTMLResponse)
async def time_budgets_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
        budgets = conn.execute(
            "SELECT category_id, weekly_hours FROM time_budgets WHERE profile_id=?", (pid,)
        ).fetchall()
    budget_map = {b["category_id"]: b["weekly_hours"] for b in budgets}
    cats = [dict(c) | {"budget_hours": budget_map.get(c["id"], 0)} for c in categories]
    return render(request, "settings.html", {
        "page": "settings",
        "categories": [dict(c) for c in categories],
        "time_budget_cats": cats,
        "show_time_budgets": True,
        "gcal_connected": False,
        "gcal_calendar_id": "primary",
        "gcal_configured": False,
        "current_bg_type": "none",
        "current_bg_preset": "",
        "current_bg_opacity": 0.7,
    })


@app.post("/settings/time-budgets", response_class=HTMLResponse)
async def save_time_budgets(request: Request):
    pid = get_profile_id(request)
    form = await request.form()
    cat_ids = form.getlist("category_id")
    hours_list = form.getlist("weekly_hours")
    with get_db() as conn:
        conn.execute("DELETE FROM time_budgets WHERE profile_id=?", (pid,))
        for cid, hrs in zip(cat_ids, hours_list):
            h = float(hrs or 0)
            if h > 0:
                conn.execute(
                    "INSERT INTO time_budgets (profile_id, category_id, weekly_hours) VALUES (?, ?, ?)",
                    (pid, int(cid), h),
                )
    return redirect(request, "/settings/time-budgets")


# ── Routes: Backup & Restore ──
@app.get("/settings/backup")
async def download_backup(request: Request):
    """Download full data backup as ZIP."""
    get_profile_id(request)
    try:
        with get_db() as conn:
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    except Exception:
        pass
    data_dir = BASE_DIR / "data"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(data_dir):
            dirs[:] = [d for d in dirs if d != "backup_before_restore"]
            for f in files:
                fp = Path(root) / f
                if fp.suffix in ('.db-wal', '.db-shm'):
                    continue
                arcname = str(fp.relative_to(data_dir))
                zf.write(fp, arcname)
    buf.seek(0)
    today = date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=work_backup_{today}.zip"},
    )


@app.post("/settings/restore")
async def restore_backup(request: Request, file: UploadFile = File(...)):
    """Restore data from uploaded ZIP backup."""
    get_profile_id(request)
    if not file.filename or not file.filename.endswith('.zip'):
        raise HTTPException(400, "ZIP 파일만 지원합니다")
    content = await file.read()
    if len(content) > 100 * 1024 * 1024:
        raise HTTPException(400, "파일이 너무 큽니다 (100MB 제한)")
    buf = io.BytesIO(content)
    try:
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
            if not any(n.endswith('.db') for n in names):
                raise HTTPException(400, "유효한 백업 파일이 아닙니다")
            for name in names:
                if name.startswith('/') or '..' in name:
                    raise HTTPException(400, "유효한 백업 파일이 아닙니다")
            data_dir = BASE_DIR / "data"
            backup_dir = data_dir / "backup_before_restore"
            if backup_dir.exists():
                shutil.rmtree(backup_dir)
            backup_dir.mkdir(exist_ok=True)
            for item in data_dir.iterdir():
                if item.name == "backup_before_restore":
                    continue
                dest = backup_dir / item.name
                if item.is_file():
                    shutil.copy2(item, dest)
                elif item.is_dir():
                    shutil.copytree(item, dest, dirs_exist_ok=True)
            zf.extractall(str(data_dir))
    except zipfile.BadZipFile:
        raise HTTPException(400, "유효한 ZIP 파일이 아닙니다")
    return RedirectResponse("/settings", status_code=303)


# ── Routes: Google Calendar OAuth ──
@app.get("/settings/gcal/connect")
async def gcal_connect(request: Request):
    if not GCAL_CLIENT_ID:
        raise HTTPException(400, "GCAL_CLIENT_ID 환경변수가 설정되지 않았습니다")
    from urllib.parse import urlencode
    pid = get_profile_id(request)
    params = urlencode({
        "client_id": GCAL_CLIENT_ID,
        "redirect_uri": _gcal_redirect_uri(request),
        "response_type": "code",
        "scope": GCAL_SCOPES,
        "access_type": "offline",
        "prompt": "consent",
        "state": str(pid),
    })
    return RedirectResponse(f"{GCAL_AUTH_URL}?{params}")


@app.get("/settings/gcal/callback")
async def gcal_callback(request: Request, code: str = "", error: str = "", state: str = ""):
    import httpx
    logger = logging.getLogger("gcal")
    logger.info(f"[GCAL] callback: code={'yes' if code else 'NO'}, error={error}, state={state}")
    if error:
        logger.error(f"[GCAL] Google returned error: {error}")
        raise HTTPException(400, f"Google 인증 오류: {error}")
    if not code:
        logger.error("[GCAL] No code received")
        raise HTTPException(400, "Google 인증 코드가 없습니다")
    pid = int(state) if state.isdigit() else get_profile_id(request)
    redirect_uri = _gcal_redirect_uri(request)
    logger.info(f"[GCAL] token exchange: pid={pid}, redirect_uri={redirect_uri}")
    async with httpx.AsyncClient() as client:
        resp = await client.post(GCAL_TOKEN_URL, data={
            "code": code,
            "client_id": GCAL_CLIENT_ID,
            "client_secret": GCAL_CLIENT_SECRET,
            "redirect_uri": redirect_uri,
            "grant_type": "authorization_code",
        })
    if resp.status_code != 200:
        logger.error(f"[GCAL] token exchange failed: {resp.status_code} {resp.text[:300]}")
        raise HTTPException(400, f"Google 토큰 교환 실패: {resp.text[:200]}")
    data = resp.json()
    logger.info(f"[GCAL] token received, expires_in={data.get('expires_in')}")
    expiry = (datetime.now() + timedelta(seconds=data["expires_in"])).isoformat()
    with get_db() as conn:
        conn.execute("""
            INSERT OR REPLACE INTO gcal_tokens (profile_id, access_token, refresh_token, token_expiry)
            VALUES (?, ?, ?, ?)
        """, (pid, data["access_token"], data.get("refresh_token", ""), expiry))
    return RedirectResponse("/settings", status_code=303)


@app.post("/settings/gcal/disconnect")
async def gcal_disconnect(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM gcal_tokens WHERE profile_id=?", (pid,))
    return RedirectResponse("/settings", status_code=303)


@app.post("/settings/gcal/calendar-id")
async def gcal_set_calendar_id(request: Request, calendar_id: str = Form("primary")):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("UPDATE gcal_tokens SET calendar_id=? WHERE profile_id=?", (calendar_id, pid))
    return RedirectResponse("/settings", status_code=303)


@app.get("/api/gcal/calendars")
async def gcal_list_calendars(request: Request):
    import httpx
    pid = get_profile_id(request)
    token = await _gcal_refresh_token(pid)
    if not token:
        return JSONResponse([])
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            f"{GCAL_API_BASE}/users/me/calendarList",
            headers={"Authorization": f"Bearer {token}"},
        )
    if resp.status_code != 200:
        return JSONResponse([])
    items = resp.json().get("items", [])
    return JSONResponse([{"id": c["id"], "summary": c.get("summary", c["id"])} for c in items])


# ── Routes: Background Settings ──
ALLOWED_BG_PRESETS = {"gradient-1", "gradient-2", "gradient-3", "gradient-4", "gradient-5", "gradient-6"}
ALLOWED_BG_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_BG_SIZE = 5 * 1024 * 1024  # 5MB


@app.post("/settings/background", response_class=HTMLResponse)
async def save_background_setting(
    request: Request,
    type: str = Form("none"),
    preset: str = Form(""),
    opacity: float = Form(0.7),
    file: UploadFile = File(None),
):
    pid = get_profile_id(request)
    bg_type = type if type in ("preset", "upload", "none") else "none"
    opacity = max(0.3, min(0.95, opacity))

    with get_db() as conn:
        set_user_setting(conn, pid, "bg_type", bg_type)
        set_user_setting(conn, pid, "bg_opacity", str(opacity))

        if bg_type == "preset" and preset in ALLOWED_BG_PRESETS:
            set_user_setting(conn, pid, "bg_preset", preset)
        elif bg_type == "upload" and file and file.filename:
            ext = Path(file.filename).suffix.lower()
            if ext not in ALLOWED_BG_EXTENSIONS:
                raise HTTPException(400, "허용되지 않는 파일 형식입니다 (jpg, png, webp만 가능)")
            content = await file.read()
            if len(content) > MAX_BG_SIZE:
                raise HTTPException(400, "파일 크기가 5MB를 초과합니다")
            filename = f"bg_{pid}_{uuid.uuid4().hex[:8]}{ext}"
            filepath = BG_DIR / filename
            filepath.write_bytes(content)
            old_image = get_user_setting(conn, pid, "bg_image", "")
            if old_image:
                old_path = BG_DIR / Path(old_image).name
                if old_path.exists():
                    old_path.unlink()
            set_user_setting(conn, pid, "bg_image", filename)
        elif bg_type == "none":
            set_user_setting(conn, pid, "bg_preset", "")

    return redirect(request, "/settings")


@app.get("/api/settings/background")
async def get_background_setting(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        bg_type = get_user_setting(conn, pid, "bg_type", "none")
        bg_preset = get_user_setting(conn, pid, "bg_preset", "")
        bg_image = get_user_setting(conn, pid, "bg_image", "")
        bg_opacity = get_user_setting(conn, pid, "bg_opacity", "0.7")
    return JSONResponse({
        "type": bg_type,
        "preset": bg_preset,
        "image": bg_image,
        "opacity": float(bg_opacity),
    })


# ── Routes: Work Logs ──
@app.get("/worklogs", response_class=HTMLResponse)
async def worklogs_page(request: Request,
                        date_param: str = Query(None, alias="date"),
                        start: str = Query(None),
                        end: str = Query(None),
                        category_id: str = Query(None, alias="cat")):
    today = date.today()

    with get_db() as conn:
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
        cats_list = [dict(c) for c in categories]

        # Range mode: start & end both provided
        start_date = validate_date_str(start) if start else None
        end_date = validate_date_str(end) if end else None
        range_mode = bool(start_date and end_date)

        if range_mode:
            # Ensure start <= end
            if start_date > end_date:
                start_date, end_date = end_date, start_date

            cat_filter = int(category_id) if category_id else None
            if cat_filter:
                logs = conn.execute("""
                    SELECT wl.*, c.name as category_name, c.color as category_color
                    FROM work_logs wl LEFT JOIN categories c ON wl.category_id = c.id
                    WHERE wl.log_date BETWEEN ? AND ? AND wl.category_id = ?
                    ORDER BY wl.log_date DESC, wl.created_at DESC
                """, (start_date, end_date, cat_filter)).fetchall()
            else:
                logs = conn.execute("""
                    SELECT wl.*, c.name as category_name, c.color as category_color
                    FROM work_logs wl LEFT JOIN categories c ON wl.category_id = c.id
                    WHERE wl.log_date BETWEEN ? AND ?
                    ORDER BY wl.log_date DESC, wl.created_at DESC
                """, (start_date, end_date)).fetchall()

            logs_list = [dict(l) for l in logs]
            total_hours = sum(l.get("hours", 0) or 0 for l in logs_list)

            # Group by date
            logs_by_date: dict = {}
            for log in logs_list:
                d = log.get("log_date", "")
                logs_by_date.setdefault(d, []).append(log)

            # For single-date nav, use today as default
            current_date = today.isoformat()
            current_dt = today
        else:
            if date_param:
                current_date = validate_date_str(date_param) or today.isoformat()
            else:
                current_date = today.isoformat()

            try:
                current_dt = datetime.strptime(current_date, "%Y-%m-%d").date()
            except ValueError:
                current_dt = today

            logs = conn.execute("""
                SELECT wl.*, c.name as category_name, c.color as category_color
                FROM work_logs wl LEFT JOIN categories c ON wl.category_id = c.id
                WHERE wl.log_date = ?
                ORDER BY wl.created_at DESC
            """, (current_date,)).fetchall()

            logs_list = [dict(l) for l in logs]
            total_hours = sum(l.get("hours", 0) or 0 for l in logs_list)
            logs_by_date = {}

    prev_date = (current_dt - timedelta(days=1)).isoformat()
    next_date = (current_dt + timedelta(days=1)).isoformat()
    is_today = current_dt == today

    return render(request, "worklogs.html", {
        "page": "worklogs",
        "logs": logs_list,
        "categories": cats_list,
        "current_date": current_date,
        "prev_date": prev_date,
        "next_date": next_date,
        "is_today": is_today,
        "total_hours": total_hours,
        "range_mode": range_mode,
        "start_date": start_date or "",
        "end_date": end_date or "",
        "logs_by_date": logs_by_date,
        "selected_category": category_id or "",
    })


@app.post("/worklogs", response_class=HTMLResponse)
async def create_worklog(request: Request,
                         title: str = Form(...),
                         content: str = Form(""),
                         hours: float = Form(0),
                         category_id: str = Form(""),
                         log_date: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    content = clamp_text(fix_mojibake(content), 5000)
    log_date = validate_date_str(log_date) or date.today().isoformat()
    cat_id = int(category_id) if category_id else None
    hours = max(0.0, min(24.0, hours))

    with get_db() as conn:
        conn.execute("""
            INSERT INTO work_logs (log_date, title, content, hours, category_id)
            VALUES (?, ?, ?, ?, ?)
        """, (log_date, title, content, hours, cat_id))

    return redirect(request, f"/worklogs?date={log_date}")


# MAGIC_BYTES and _check_image_magic imported from common.image

@app.post("/worklogs/upload-image")
async def upload_worklog_image(request: Request, file: UploadFile = File(...)):
    pid = get_profile_id(request)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(400, detail="파일 크기가 10MB를 초과합니다")

    ext = Path(file.filename or "img.png").suffix.lower()
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        raise HTTPException(400, detail="지원하지 않는 이미지 형식입니다")

    if not _check_image_magic(content, ext):
        raise HTTPException(400, detail="파일 내용이 이미지가 아닙니다")

    filename = f"wl_{pid}_{uuid.uuid4().hex[:8]}{ext}"
    (WORKLOG_IMG_DIR / filename).write_bytes(content)

    return JSONResponse({"url": f"/worklog-images/{filename}"})


@app.get("/worklogs/{log_id}/edit", response_class=HTMLResponse)
async def edit_worklog_form(request: Request, log_id: int):
    with get_db() as conn:
        log = conn.execute(
            "SELECT wl.*, c.name as category_name, c.color as category_color "
            "FROM work_logs wl LEFT JOIN categories c ON wl.category_id = c.id WHERE wl.id=?",
            (log_id,),
        ).fetchone()
        if not log:
            raise HTTPException(404)
        categories = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
    return templates.TemplateResponse(request, "partials/worklog_edit_form.html", {
        "log": dict(log), "categories": [dict(c) for c in categories],
    })


@app.put("/worklogs/{log_id}", response_class=HTMLResponse)
async def update_worklog(request: Request, log_id: int,
                         title: str = Form(...),
                         content: str = Form(""),
                         hours: float = Form(0),
                         category_id: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    content = clamp_text(fix_mojibake(content), 5000)
    cat_id = int(category_id) if category_id else None
    hours = max(0.0, min(24.0, hours))

    with get_db() as conn:
        log = conn.execute("SELECT log_date FROM work_logs WHERE id=?", (log_id,)).fetchone()
        conn.execute("""
            UPDATE work_logs SET title=?, content=?, hours=?, category_id=?,
                   updated_at=datetime('now','localtime')
            WHERE id=?
        """, (title, content, hours, cat_id, log_id))

    log_date = log["log_date"] if log else date.today().isoformat()
    return redirect(request, f"/worklogs?date={log_date}")


@app.delete("/worklogs/{log_id}", response_class=HTMLResponse)
async def delete_worklog(request: Request, log_id: int):
    with get_db() as conn:
        row = conn.execute("SELECT content FROM work_logs WHERE id=?", (log_id,)).fetchone()
        if row and row["content"]:
            import re as _re
            for img_path in _re.findall(r'!\[[^\]]*\]\((/worklog-images/([^)]+))\)', row["content"]):
                img_file = WORKLOG_IMG_DIR / img_path[1]
                if img_file.is_file():
                    img_file.unlink()
        conn.execute("DELETE FROM work_logs WHERE id=?", (log_id,))
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/worklogs")


# ── Routes: Notices ──
@app.get("/notices", response_class=HTMLResponse)
async def notices_page(request: Request):
    with get_db() as conn:
        notices = conn.execute("""
            SELECT * FROM notices
            ORDER BY pinned DESC, created_at DESC
        """).fetchall()
    return render(request, "notices.html", {
        "page": "notices",
        "notices": [dict(n) for n in notices],
    })


@app.post("/notices", response_class=HTMLResponse)
async def create_notice(request: Request,
                        title: str = Form(...),
                        content: str = Form(""),
                        priority: int = Form(0)):
    title = clamp_text(fix_mojibake(title), 200)
    content = clamp_text(fix_mojibake(content), 5000)
    priority = 1 if priority else 0

    with get_db() as conn:
        conn.execute("""
            INSERT INTO notices (title, content, priority)
            VALUES (?, ?, ?)
        """, (title, content, priority))

    return redirect(request, "/notices")


@app.get("/notices/{notice_id}/edit", response_class=HTMLResponse)
async def edit_notice_form(request: Request, notice_id: int):
    with get_db() as conn:
        notice = conn.execute("SELECT * FROM notices WHERE id=?", (notice_id,)).fetchone()
        if not notice:
            raise HTTPException(404)
    return templates.TemplateResponse(request, "partials/notice_edit_form.html", {
        "notice": dict(notice),
    })


@app.put("/notices/{notice_id}", response_class=HTMLResponse)
async def update_notice(request: Request, notice_id: int,
                        title: str = Form(...),
                        content: str = Form(""),
                        priority: int = Form(0)):
    title = clamp_text(fix_mojibake(title), 200)
    content = clamp_text(fix_mojibake(content), 5000)
    priority = 1 if priority else 0

    with get_db() as conn:
        conn.execute("""
            UPDATE notices SET title=?, content=?, priority=?,
                   updated_at=datetime('now','localtime')
            WHERE id=?
        """, (title, content, priority, notice_id))

    return redirect(request, "/notices")


@app.delete("/notices/{notice_id}", response_class=HTMLResponse)
async def delete_notice(request: Request, notice_id: int):
    with get_db() as conn:
        conn.execute("DELETE FROM notices WHERE id=?", (notice_id,))
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/notices")


@app.post("/notices/{notice_id}/pin", response_class=HTMLResponse)
async def toggle_pin_notice(request: Request, notice_id: int):
    with get_db() as conn:
        notice = conn.execute("SELECT pinned FROM notices WHERE id=?", (notice_id,)).fetchone()
        if not notice:
            raise HTTPException(404)
        new_pinned = 0 if notice["pinned"] else 1
        conn.execute("UPDATE notices SET pinned=? WHERE id=?", (new_pinned, notice_id))
    return redirect(request, "/notices")




# ── Routes: Form Templates (양식) ──
@app.get("/forms", response_class=HTMLResponse)
async def forms_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpls = conn.execute(
            "SELECT * FROM form_templates WHERE profile_id IN (?, 0) ORDER BY profile_id ASC, updated_at DESC", (pid,)
        ).fetchall()
        entry_counts = {}
        field_counts = {}
        for t in tpls:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM form_entries WHERE template_id=? AND profile_id=?",
                (t["id"], pid),
            ).fetchone()[0]
            entry_counts[t["id"]] = cnt
            try:
                field_counts[t["id"]] = len(json.loads(t["fields"]))
            except (json.JSONDecodeError, TypeError):
                field_counts[t["id"]] = 0
    return render(request, "forms.html", {
        "page": "forms",
        "templates": [dict(t) for t in tpls],
        "entry_counts": entry_counts,
        "field_counts": field_counts,
    })


@app.get("/forms/new", response_class=HTMLResponse)
async def form_builder_new(request: Request):
    return render(request, "form_builder.html", {
        "page": "forms",
        "mode": "create",
        "tpl": None,
    })


# _parse_excel_with_merges and _infer_field_type imported from common.excel


@app.post("/forms/upload", response_class=HTMLResponse)
async def create_form_from_file(request: Request):
    import csv, io
    pid = get_profile_id(request)
    form = await request.form()
    file = form.get("file")
    if not file or not hasattr(file, "filename") or not file.filename:
        raise HTTPException(400, "파일을 선택해주세요")

    fname = file.filename.lower()
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "파일 크기 초과 (최대 10MB)")

    headers = []
    sample_rows = []

    if fname.endswith((".xlsx", ".xls")):
        headers, sample_rows = _parse_excel_with_merges(content)
    elif fname.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        header_row = next(reader, None)
        if header_row:
            headers = [h.strip() if h.strip() else f"열{i+1}" for i, h in enumerate(header_row)]
        for i, row in enumerate(reader):
            if i >= 500:
                break
            sample_rows.append(row)
    else:
        raise HTTPException(400, "xlsx 또는 csv 파일만 지원합니다")

    if not headers:
        raise HTTPException(400, "헤더를 찾을 수 없습니다")

    fields = []
    for i, h in enumerate(headers):
        col_vals = [r[i] for r in sample_rows if i < len(r)]
        ftype = _infer_field_type(col_vals)
        fields.append({"label": h, "type": ftype, "required": False})

    tpl_name = Path(file.filename).stem
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields) VALUES (?, ?, ?, ?)",
            (pid, tpl_name, f"파일 업로드로 생성 ({file.filename})", json.dumps(fields, ensure_ascii=False)),
        )
        tpl_id = cur.lastrowid

        for row in sample_rows:
            data = {}
            for j, f in enumerate(fields):
                if j < len(row) and row[j] is not None:
                    val = row[j]
                    if hasattr(val, "strftime"):
                        from datetime import datetime as _dt, time as _tm
                        if isinstance(val, _tm):
                            val = val.strftime("%H:%M")
                        elif isinstance(val, _dt):
                            if val.year == 1900 and val.month == 1 and val.day == 1:
                                val = val.strftime("%H:%M")
                            else:
                                val = val.strftime("%Y-%m-%d")
                        else:
                            val = val.strftime("%Y-%m-%d")
                    data[f["label"]] = str(val) if val else ""
                else:
                    data[f["label"]] = ""
            entry_date_val = data.get(headers[0], "") if fields[0]["type"] == "date" else date.today().isoformat()
            if not validate_date_str(entry_date_val):
                entry_date_val = date.today().isoformat()
            conn.execute(
                "INSERT INTO form_entries (template_id, profile_id, entry_date, values_json) VALUES (?, ?, ?, ?)",
                (tpl_id, pid, entry_date_val, json.dumps(data, ensure_ascii=False)),
            )

    return redirect(request, f"/forms/{tpl_id}/entries")


@app.post("/forms", response_class=HTMLResponse)
async def create_form_template(request: Request):
    pid = get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    description = clamp_text(fix_mojibake(form.get("description", "")), 500)
    emoji = clamp_text(fix_mojibake(form.get("emoji", "")), 2) or "📝"
    color = form.get("color", "#6366f1") or "#6366f1"
    if not name:
        raise HTTPException(400, "양식 이름은 필수입니다")
    fields = []
    idx = 0
    while True:
        label = form.get(f"field_{idx}_label")
        if label is None:
            break
        ftype = form.get(f"field_{idx}_type", "text")
        required = form.get(f"field_{idx}_required") == "on"
        options = form.get(f"field_{idx}_options", "")
        default_val = clamp_text(fix_mojibake(form.get(f"field_{idx}_default", "")), 500)
        copy_prev = form.get(f"field_{idx}_copy_prev") == "on"
        field = {
            "label": clamp_text(fix_mojibake(label), 100),
            "type": ftype if ftype in ("text", "textarea", "number", "dropdown", "date", "checkbox", "table") else "text",
            "required": required,
            "default": default_val,
            "copy_prev": copy_prev,
        }
        if ftype == "dropdown" and options:
            field["options"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        if ftype == "table" and options:
            field["columns"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        fields.append(field)
        idx += 1
    if not fields:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color) VALUES (?, ?, ?, ?, ?, ?)",
            (pid, name, description, json.dumps(fields, ensure_ascii=False), emoji, color),
        )
    event_bus.broadcast("forms")
    return redirect(request, "/forms")


@app.get("/forms/{tpl_id}/edit", response_class=HTMLResponse)
async def form_builder_edit(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
    tpl_dict = dict(tpl)
    tpl_dict["fields"] = json.loads(tpl_dict["fields"])
    return render(request, "form_builder.html", {
        "page": "forms",
        "mode": "edit",
        "tpl": tpl_dict,
    })


@app.put("/forms/{tpl_id}", response_class=HTMLResponse)
async def update_form_template(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    description = clamp_text(fix_mojibake(form.get("description", "")), 500)
    emoji = clamp_text(fix_mojibake(form.get("emoji", "")), 2) or "📝"
    color = form.get("color", "#6366f1") or "#6366f1"
    if not name:
        raise HTTPException(400, "양식 이름은 필수입니다")
    fields = []
    idx = 0
    while True:
        label = form.get(f"field_{idx}_label")
        if label is None:
            break
        ftype = form.get(f"field_{idx}_type", "text")
        required = form.get(f"field_{idx}_required") == "on"
        options = form.get(f"field_{idx}_options", "")
        default_val = clamp_text(fix_mojibake(form.get(f"field_{idx}_default", "")), 500)
        copy_prev = form.get(f"field_{idx}_copy_prev") == "on"
        field = {
            "label": clamp_text(fix_mojibake(label), 100),
            "type": ftype if ftype in ("text", "textarea", "number", "dropdown", "date", "checkbox", "table") else "text",
            "required": required,
            "default": default_val,
            "copy_prev": copy_prev,
        }
        if ftype == "dropdown" and options:
            field["options"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        if ftype == "table" and options:
            field["columns"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        fields.append(field)
        idx += 1
    if not fields:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    with get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM form_templates WHERE id=? AND profile_id=?", (tpl_id, pid)
        ).fetchone()
        if not existing:
            raise HTTPException(404)
        conn.execute(
            "UPDATE form_templates SET name=?, description=?, fields=?, emoji=?, color=?, updated_at=datetime('now','localtime') WHERE id=? AND profile_id=?",
            (name, description, json.dumps(fields, ensure_ascii=False), emoji, color, tpl_id, pid),
        )
    event_bus.broadcast("forms")
    return redirect(request, "/forms")


@app.post("/forms/{tpl_id}/clone", response_class=HTMLResponse)
async def clone_form_template(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color) VALUES (?, ?, ?, ?, ?, ?)",
            (pid, f"{tpl['name']} (복사)", tpl["description"], tpl["fields"], tpl["emoji"], tpl["color"]),
        )
    event_bus.broadcast("forms")
    return redirect(request, "/forms")


@app.delete("/forms/{tpl_id}", response_class=HTMLResponse)
async def delete_form_template(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM form_templates WHERE id=? AND profile_id=?", (tpl_id, pid))
    event_bus.broadcast("forms")
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, "/forms")


@app.get("/forms/{form_id}/export-json")
async def export_form_json(request: Request, form_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (form_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
    fields = json.loads(tpl["fields"])
    data = {
        "name": tpl["name"],
        "description": tpl["description"] or "",
        "emoji": tpl["emoji"] or "📝",
        "color": tpl["color"] or "#6366f1",
        "fields": fields,
    }
    content = json.dumps(data, ensure_ascii=False, indent=2)
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{tpl["name"]}.json"'},
    )


@app.post("/forms/import-json")
async def import_form_json(request: Request, file: UploadFile = File(...)):
    pid = get_profile_id(request)
    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(400, "유효한 JSON 파일이 아닙니다")
    if not isinstance(data, dict) or "name" not in data or "fields" not in data:
        raise HTTPException(400, "필수 항목(name, fields)이 없습니다")
    if not isinstance(data["fields"], list) or len(data["fields"]) == 0:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    name = clamp_text(str(data["name"]), 100)
    description = clamp_text(str(data.get("description", "")), 500)
    emoji = str(data.get("emoji", "📝"))[:2] or "📝"
    color = str(data.get("color", "#6366f1"))[:20] or "#6366f1"
    valid_types = {"text", "textarea", "number", "dropdown", "date", "checkbox", "table"}
    fields = []
    for f in data["fields"]:
        if not isinstance(f, dict) or "label" not in f:
            continue
        field = {
            "label": clamp_text(str(f["label"]), 100),
            "type": f.get("type", "text") if f.get("type") in valid_types else "text",
            "required": bool(f.get("required", False)),
        }
        if field["type"] == "dropdown" and f.get("options"):
            field["options"] = [str(o) for o in f["options"]]
        if field["type"] == "table" and f.get("columns"):
            field["columns"] = [str(c) for c in f["columns"]]
        if f.get("default"):
            field["default"] = str(f["default"])
        if f.get("copy_prev"):
            field["copy_prev"] = True
        fields.append(field)
    if not fields:
        raise HTTPException(400, "유효한 필드가 없습니다")
    with get_db() as conn:
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color) VALUES (?, ?, ?, ?, ?, ?)",
            (pid, name, description, json.dumps(fields, ensure_ascii=False), emoji, color),
        )
    event_bus.broadcast("forms")
    return redirect(request, "/forms")


# ── Routes: Form Entries (양식 작성) ──
@app.get("/forms/{tpl_id}/entries", response_class=HTMLResponse)
async def form_entries_page(request: Request, tpl_id: int, date_param: str = Query(None, alias="date")):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        tpl_dict = dict(tpl)
        tpl_dict["fields"] = json.loads(tpl_dict["fields"])

        today_str = date_param if validate_date_str(date_param or "") else date.today().isoformat()
        entries = conn.execute(
            "SELECT * FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date=? ORDER BY created_at DESC",
            (tpl_id, pid, today_str),
        ).fetchall()
        parsed_entries = []
        for e in entries:
            ed = dict(e)
            ed["data"] = json.loads(ed["values_json"])
            parsed_entries.append(ed)

        # Fetch yesterday's last entry for copy_prev defaults
        yesterday = (datetime.strptime(today_str, "%Y-%m-%d").date() - timedelta(days=1)).isoformat()
        prev_entry = conn.execute(
            "SELECT values_json FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date=? ORDER BY created_at DESC LIMIT 1",
            (tpl_id, pid, yesterday),
        ).fetchone()
        prev_values = json.loads(prev_entry["values_json"]) if prev_entry else {}

    current_date = datetime.strptime(today_str, "%Y-%m-%d").date()
    prev_date = (current_date - timedelta(days=1)).isoformat()
    next_date = (current_date + timedelta(days=1)).isoformat()

    # Build field defaults: copy_prev overrides static default
    field_defaults = {}
    for f in tpl_dict["fields"]:
        if f.get("copy_prev") and f["label"] in prev_values:
            field_defaults[f["label"]] = prev_values[f["label"]]
        elif f.get("default"):
            field_defaults[f["label"]] = f["default"]

    return render(request, "form_entries.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entries": parsed_entries,
        "current_date": today_str,
        "prev_date": prev_date,
        "next_date": next_date,
        "is_today": today_str == date.today().isoformat(),
        "field_defaults": field_defaults,
    })


@app.get("/forms/{tpl_id}/entries/new", response_class=HTMLResponse)
async def form_entry_new(request: Request, tpl_id: int, date_param: str = Query(None, alias="date")):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        tpl_dict = dict(tpl)
        tpl_dict["fields"] = json.loads(tpl_dict["fields"])
        entry_date = date_param if validate_date_str(date_param or "") else date.today().isoformat()

        # Fetch yesterday's last entry for copy_prev defaults
        yesterday = (datetime.strptime(entry_date, "%Y-%m-%d").date() - timedelta(days=1)).isoformat()
        prev_entry = conn.execute(
            "SELECT values_json FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date=? ORDER BY created_at DESC LIMIT 1",
            (tpl_id, pid, yesterday),
        ).fetchone()
        prev_values = json.loads(prev_entry["values_json"]) if prev_entry else {}

    field_defaults = {}
    for f in tpl_dict["fields"]:
        if f.get("copy_prev") and f["label"] in prev_values:
            field_defaults[f["label"]] = prev_values[f["label"]]
        elif f.get("default"):
            field_defaults[f["label"]] = f["default"]

    return render(request, "form_entry_edit.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entry": None,
        "mode": "create",
        "current_date": entry_date,
        "field_defaults": field_defaults,
    })


@app.post("/forms/{tpl_id}/entries", response_class=HTMLResponse)
async def create_form_entry(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    form = await request.form()
    entry_date = validate_date_str(form.get("entry_date", "")) or date.today().isoformat()
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        fields = json.loads(tpl["fields"])
        values = {}
        for i, f in enumerate(fields):
            key = f"field_{i}"
            if f["type"] == "checkbox":
                values[f["label"]] = form.get(key) == "on"
            elif f["type"] == "number":
                try:
                    nv = float(form.get(key, 0) or 0)
                    values[f["label"]] = int(nv) if nv == int(nv) else nv
                except (ValueError, TypeError):
                    values[f["label"]] = 0
            elif f["type"] == "table":
                try:
                    values[f["label"]] = json.loads(form.get(key, "[]"))
                except (json.JSONDecodeError, TypeError):
                    values[f["label"]] = []
            else:
                val = fix_mojibake(form.get(key, ""))
                values[f["label"]] = clamp_text(val, 5000)
        cur = conn.execute(
            "INSERT INTO form_entries (template_id, profile_id, entry_date, values_json) VALUES (?, ?, ?, ?)",
            (tpl_id, pid, entry_date, json.dumps(values, ensure_ascii=False)),
        )
        new_id = cur.lastrowid
    if request.headers.get("X-Inline"):
        return JSONResponse({"id": new_id})
    event_bus.broadcast("forms")
    return redirect(request, f"/forms/{tpl_id}/entries?date={entry_date}")


@app.get("/forms/{tpl_id}/entries/{entry_id}/edit", response_class=HTMLResponse)
async def form_entry_edit(request: Request, tpl_id: int, entry_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        entry = conn.execute(
            "SELECT * FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        ).fetchone()
        if not entry:
            raise HTTPException(404)
    tpl_dict = dict(tpl)
    tpl_dict["fields"] = json.loads(tpl_dict["fields"])
    entry_dict = dict(entry)
    entry_dict["data"] = json.loads(entry_dict["values_json"])
    return render(request, "form_entry_edit.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entry": entry_dict,
        "mode": "edit",
        "field_defaults": {},
    })


@app.put("/forms/{tpl_id}/entries/{entry_id}", response_class=HTMLResponse)
async def update_form_entry(request: Request, tpl_id: int, entry_id: int):
    pid = get_profile_id(request)
    form = await request.form()
    with get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        entry = conn.execute(
            "SELECT * FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        ).fetchone()
        if not entry:
            raise HTTPException(404)
        fields = json.loads(tpl["fields"])
        values = {}
        for i, f in enumerate(fields):
            key = f"field_{i}"
            if f["type"] == "checkbox":
                values[f["label"]] = form.get(key) == "on"
            elif f["type"] == "number":
                try:
                    nv = float(form.get(key, 0) or 0)
                    values[f["label"]] = int(nv) if nv == int(nv) else nv
                except (ValueError, TypeError):
                    values[f["label"]] = 0
            elif f["type"] == "table":
                try:
                    values[f["label"]] = json.loads(form.get(key, "[]"))
                except (json.JSONDecodeError, TypeError):
                    values[f["label"]] = []
            else:
                val = fix_mojibake(form.get(key, ""))
                values[f["label"]] = clamp_text(val, 5000)
        conn.execute(
            "UPDATE form_entries SET values_json=?, updated_at=datetime('now','localtime') WHERE id=? AND profile_id=?",
            (json.dumps(values, ensure_ascii=False), entry_id, pid),
        )
    if request.headers.get("X-Inline"):
        return JSONResponse({"ok": True})
    event_bus.broadcast("forms")
    return redirect(request, f"/forms/{tpl_id}/entries?date={entry['entry_date']}")


@app.delete("/forms/{tpl_id}/entries/{entry_id}", response_class=HTMLResponse)
async def delete_form_entry(request: Request, tpl_id: int, entry_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute(
            "DELETE FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        )
    event_bus.broadcast("forms")
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return redirect(request, f"/forms/{tpl_id}/entries")


def _collect_export_data(conn, tpl_id, pid, date=None):
    tpl = conn.execute("SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)).fetchone()
    if not tpl:
        return None, None, None
    fields = json.loads(tpl["fields"])
    where = "template_id=? AND profile_id=?"
    params = [tpl_id, pid]
    if date:
        where += " AND entry_date=?"
        params.append(date)
    entries = conn.execute(f"SELECT * FROM form_entries WHERE {where} ORDER BY entry_date, id", params).fetchall()
    return tpl, fields, entries


@app.get("/forms/{tpl_id}/entries/export")
async def export_entries(request: Request, tpl_id: int, date: Optional[str] = None, fmt: str = "xlsx"):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl, fields, entries = _collect_export_data(conn, tpl_id, pid, date)
    if not tpl:
        raise HTTPException(404)

    non_table_fields = [f for f in fields if f["type"] != "table"]
    has_date_field = any(f["type"] == "date" for f in non_table_fields)
    if has_date_field:
        headers_list = [f["label"] for f in non_table_fields]
    else:
        headers_list = ["날짜"] + [f["label"] for f in non_table_fields]
    rows = []
    for e in entries:
        data = json.loads(e["values_json"])
        if has_date_field:
            row = []
        else:
            row = [e["entry_date"]]
        for f in non_table_fields:
            val = data.get(f["label"], "")
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row.append(val)
        rows.append(row)

    if fmt == "csv":
        import csv, io as _io
        buf = _io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers_list)
        for r in rows:
            writer.writerow(r)
        content = buf.getvalue()
        fname = f"{tpl['name']}_{date or 'all'}.csv"
        return Response(content=content.encode("utf-8-sig"), media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tpl["name"][:31]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for ci, h in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border

    for ci in range(1, len(headers_list) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 15

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{tpl['name']}_{date or 'all'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Routes: Plans (weekly/monthly) ──


@app.get("/plans", response_class=HTMLResponse)
async def plans_redirect(request: Request, view: str = "week", offset: int = 0):
    return redirect(request, f"/?plan_view={view}&plan_offset={offset}")


# ── Quick-add from dashboard ──
@app.post("/quick-todo", response_class=HTMLResponse)
async def quick_add_todo(request: Request,
                         title: str = Form(...),
                         due_date: str = Form("")):
    title = clamp_text(fix_mojibake(title), 200)
    due_date = validate_date_str(due_date) or date.today().isoformat()
    pid = get_profile_id(request)
    with get_db() as conn:
        max_order = conn.execute("SELECT COALESCE(MAX(sort_order),0) FROM todos").fetchone()[0]
        conn.execute("""
            INSERT INTO todos (title, due_date, sort_order, profile_id) VALUES (?, ?, ?, ?)
        """, (title, due_date, max_order + 1, pid))
    return redirect(request, "/")




# ── Routes: Memo view partial ──
@app.get("/memos/{memo_id}/view", response_class=HTMLResponse)
async def memo_view(request: Request, memo_id: int):
    with get_db() as conn:
        memo = conn.execute(
            "SELECT m.*, c.name as category_name, c.color as category_color "
            "FROM memos m LEFT JOIN categories c ON m.category_id = c.id WHERE m.id=?",
            (memo_id,),
        ).fetchone()
        if not memo:
            raise HTTPException(404)
    return templates.TemplateResponse(request, "partials/memo_card.html", {
        "memo": dict(memo),
    })


# ── iCal Feed (Calendar Subscription) ──
@app.get("/cal/{profile_id}.ics")
async def ical_feed(request: Request, profile_id: int):
    pid = profile_id
    with get_db() as conn:
        profile = conn.execute("SELECT * FROM work_profiles WHERE id=?", (pid,)).fetchone()
        if not profile:
            raise HTTPException(404)
        events = conn.execute(
            "SELECT * FROM events WHERE profile_id=? ORDER BY start_time", (pid,)
        ).fetchall()
        todos = conn.execute(
            "SELECT * FROM todos WHERE profile_id=? AND due_date IS NOT NULL AND due_date != ''",
            (pid,),
        ).fetchall()
        form_entries = conn.execute(
            "SELECT fe.id, fe.entry_date, fe.values_json, ft.name as tpl_name "
            "FROM form_entries fe JOIN form_templates ft ON fe.template_id=ft.id "
            "WHERE fe.profile_id=? AND fe.entry_date IS NOT NULL", (pid,)
        ).fetchall()

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//WorkPlanner//iCal//KR",
        f"X-WR-CALNAME:{profile['name']} Work Planner",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
    ]

    def esc(s):
        return str(s).replace("\\", "\\\\").replace(",", "\\,").replace(";", "\\;").replace("\n", "\\n")

    for ev in events:
        uid = f"event-{ev['id']}@workplanner"
        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        st = ev["start_time"].replace("-", "").replace(":", "").replace(" ", "T")
        if len(st) == 8:
            lines.append(f"DTSTART;VALUE=DATE:{st}")
        else:
            lines.append(f"DTSTART:{st}")
        if ev["end_time"]:
            et = ev["end_time"].replace("-", "").replace(":", "").replace(" ", "T")
            lines.append(f"DTEND:{et}")
        lines.append(f"SUMMARY:{esc(ev['title'])}")
        if ev["memo"]:
            lines.append(f"DESCRIPTION:{esc(ev['memo'])}")
        lines.append("END:VEVENT")

    for td in todos:
        uid = f"todo-{td['id']}@workplanner"
        lines.append("BEGIN:VTODO")
        lines.append(f"UID:{uid}")
        dd = td["due_date"].replace("-", "")
        lines.append(f"DUE;VALUE=DATE:{dd}")
        lines.append(f"SUMMARY:{esc(td['title'])}")
        if td["completed"]:
            lines.append("STATUS:COMPLETED")
        else:
            lines.append("STATUS:NEEDS-ACTION")
        lines.append("END:VTODO")

    for fe in form_entries:
        uid = f"form-{fe['id']}@workplanner"
        lines.append("BEGIN:VEVENT")
        lines.append(f"UID:{uid}")
        dd = fe["entry_date"].replace("-", "")
        lines.append(f"DTSTART;VALUE=DATE:{dd}")
        try:
            data = json.loads(fe["values_json"])
            summary_parts = [str(v) for v in list(data.values())[:3] if v]
            summary = f"[{fe['tpl_name']}] {' / '.join(summary_parts)}"
        except (json.JSONDecodeError, TypeError):
            summary = f"[{fe['tpl_name']}] {fe['entry_date']}"
        lines.append(f"SUMMARY:{esc(summary[:100])}")
        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")
    ical_content = "\r\n".join(lines)
    return Response(
        content=ical_content.encode("utf-8"),
        media_type="text/calendar; charset=utf-8",
        headers={"Content-Disposition": f'inline; filename="calendar_{profile_id}.ics"'},
    )


# ── Routes: Audit Log ──
@app.get("/audit-log", response_class=HTMLResponse)
async def audit_log_page(request: Request, entity_type: str = "", limit: int = 50):
    with get_db() as conn:
        if entity_type:
            rows = conn.execute(
                "SELECT * FROM audit_log WHERE entity_type=? ORDER BY created_at DESC LIMIT ?",
                (entity_type, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM audit_log ORDER BY created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
    logs = []
    for r in rows:
        d = dict(r)
        d["changes"] = json.loads(d.get("changes_json") or "{}")
        logs.append(d)
    return render(request, "audit_log.html", {
        "page": "audit-log",
        "logs": logs,
        "entity_type": entity_type,
    })


# ── Routes: Search ──
@app.get("/search", response_class=HTMLResponse)
async def search_page(request: Request, q: str = ""):
    pid = get_profile_id(request)
    results = {"todos": [], "events": [], "memos": [], "notices": [], "worklogs": [], "entries": []}
    if q and len(q) >= 2:
        like = f"%{q}%"
        with get_db() as conn:
            results["todos"] = [dict(r) for r in conn.execute(
                "SELECT id, title, due_date FROM todos WHERE profile_id=? AND (title LIKE ? OR description LIKE ?) LIMIT 20",
                (pid, like, like)).fetchall()]
            results["events"] = [dict(r) for r in conn.execute(
                "SELECT id, title, start_time FROM events WHERE profile_id=? AND (title LIKE ? OR memo LIKE ?) ORDER BY start_time DESC LIMIT 20",
                (pid, like, like)).fetchall()]
            results["memos"] = [dict(r) for r in conn.execute(
                "SELECT id, content, created_at FROM memos WHERE profile_id=? AND content LIKE ? ORDER BY created_at DESC LIMIT 20",
                (pid, like)).fetchall()]
            results["notices"] = [dict(r) for r in conn.execute(
                "SELECT id, title, created_at FROM notices WHERE title LIKE ? OR content LIKE ? ORDER BY created_at DESC LIMIT 20",
                (like, like)).fetchall()]
            results["worklogs"] = [dict(r) for r in conn.execute(
                "SELECT id, content, log_date FROM work_logs WHERE content LIKE ? ORDER BY log_date DESC LIMIT 20",
                (like,)).fetchall()]
            form_entries = conn.execute(
                "SELECT fe.id, fe.entry_date, fe.values_json, ft.name as tpl_name, ft.id as tpl_id "
                "FROM form_entries fe JOIN form_templates ft ON fe.template_id=ft.id "
                "WHERE fe.profile_id=? AND fe.values_json LIKE ? ORDER BY fe.entry_date DESC LIMIT 20",
                (pid, like)).fetchall()
            results["entries"] = [dict(r) for r in form_entries]
    total = sum(len(v) for v in results.values())
    return render(request, "search.html", {"page": "search", "q": q, "results": results, "total": total})


# ── Focus mode ──
@app.post("/focus/complete", response_class=HTMLResponse)
async def focus_complete(request: Request):
    pid = get_profile_id(request)
    body = await request.json()
    minutes = max(1, min(480, int(body.get("minutes", 25))))
    hours = round(minutes / 60, 2)
    cat_id = body.get("category_id") or None
    title = clamp_text(body.get("title", ""), 200) or f"집중 모드 {minutes}분"
    with get_db() as conn:
        conn.execute(
            "INSERT INTO work_logs (log_date, title, content, hours, category_id) VALUES (?,?,?,?,?)",
            (date.today().isoformat(), title, f"집중 모드 {minutes}분 완료", hours, cat_id))
    return JSONResponse({"ok": True, "hours": hours})


# ── Routes: D-day ──
@app.get("/ddays", response_class=HTMLResponse)
async def ddays_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM ddays WHERE profile_id=? ORDER BY target_date ASC", (pid,)
        ).fetchall()
    ddays = []
    for r in rows:
        d = dict(r)
        d["dday"] = calc_dday(d["target_date"])
        if not d.get("icon"):
            d["icon"] = "\U0001f3af"
        ddays.append(d)
    return render(request, "ddays.html", {"page": "ddays", "ddays": ddays})


@app.post("/ddays", response_class=HTMLResponse)
async def create_dday(request: Request, title: str = Form(...), target_date: str = Form(...), icon: str = Form("\U0001f3af")):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute(
            "INSERT INTO ddays (profile_id, title, target_date, icon) VALUES (?,?,?,?)",
            (pid, title, target_date, icon or "\U0001f3af"),
        )
    return redirect(request, "/ddays")


@app.delete("/ddays/{dday_id}", response_class=HTMLResponse)
async def delete_dday(request: Request, dday_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM ddays WHERE id=? AND profile_id=?", (dday_id, pid))
    return HTMLResponse("")


# ── Routes: Links ──
@app.get("/links", response_class=HTMLResponse)
async def links_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM links WHERE profile_id=? ORDER BY created_at DESC", (pid,)
        ).fetchall()
        cats = conn.execute(
            "SELECT DISTINCT category FROM links WHERE profile_id=? AND category != '' ORDER BY category",
            (pid,),
        ).fetchall()
    links = [dict(r) for r in rows]
    categories = [r["category"] for r in cats]
    # Read tunnel data for cross-planner links
    tunnel_data = None
    tunnel_file = BASE_DIR / "tunnel-url.txt"
    if tunnel_file.exists():
        try:
            import json as _json
            text = tunnel_file.read_text().strip()
            if text.startswith("{"):
                tunnel_data = _json.loads(text)
            else:
                tunnel_data = {"jm": {"url": text}}
        except Exception:
            pass
    return render(request, "links.html", {
        "page": "links",
        "links": links,
        "link_categories": categories,
        "tunnel_data": tunnel_data,
        "current_app": "jm",
    })


@app.post("/links", response_class=HTMLResponse)
async def create_link(request: Request, title: str = Form(...), url: str = Form(...),
                      category: str = Form(""), description: str = Form("")):
    if not url.startswith(('http://', 'https://', '/')):
        raise HTTPException(400, detail="유효하지 않은 URL입니다")
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute(
            "INSERT INTO links (profile_id, title, url, category, description) VALUES (?,?,?,?,?)",
            (pid, title, url, category, description),
        )
    return redirect(request, "/links")


@app.delete("/links/{link_id}", response_class=HTMLResponse)
async def delete_link(request: Request, link_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        conn.execute("DELETE FROM links WHERE id=? AND profile_id=?", (link_id, pid))
    return HTMLResponse("")


# ── Routes: Stats ──
@app.get("/stats", response_class=HTMLResponse)
async def stats_page(request: Request):
    pid = get_profile_id(request)
    with get_db() as conn:
        stats = get_stats(conn, pid)
        chart_data = get_weekly_chart_data(conn, pid)

        # Total counts
        total_all = conn.execute("SELECT COUNT(*) FROM todos WHERE profile_id=?", (pid,)).fetchone()[0]
        total_completed = conn.execute("SELECT COUNT(*) FROM todos WHERE profile_id=? AND completed=1", (pid,)).fetchone()[0]
        total_rate = round(total_completed / total_all * 100) if total_all > 0 else 0

        # Category stats
        cat_stats = []
        cats = conn.execute("SELECT * FROM categories ORDER BY sort_order").fetchall()
        for c in cats:
            total = conn.execute(
                "SELECT COUNT(*) FROM todos WHERE category_id=? AND profile_id=?", (c["id"], pid)
            ).fetchone()[0]
            done = conn.execute(
                "SELECT COUNT(*) FROM todos WHERE category_id=? AND profile_id=? AND completed=1", (c["id"], pid)
            ).fetchone()[0]
            cat_stats.append({"name": c["name"], "color": c["color"], "total": total, "done": done})

        # Monthly trend
        monthly_data = []
        today = date.today()
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m < 1:
                m += 12
                y -= 1
            label = f"{y}-{m:02d}"
            month_start = f"{y}-{m:02d}-01"
            if m == 12:
                month_end = f"{y + 1}-01-01"
            else:
                month_end = f"{y}-{m + 1:02d}-01"
            total = conn.execute(
                "SELECT COUNT(*) FROM todos WHERE profile_id=? AND created_at>=? AND created_at<?",
                (pid, month_start, month_end),
            ).fetchone()[0]
            done = conn.execute(
                "SELECT COUNT(*) FROM todos WHERE profile_id=? AND completed=1 AND completed_at>=? AND completed_at<?",
                (pid, month_start, month_end),
            ).fetchone()[0]
            monthly_data.append({"label": label, "total": total, "done": done})

        # Monthly events
        monthly_events = conn.execute("""
            SELECT strftime('%Y-%m', start_time) as m, COUNT(*) as c
            FROM events WHERE profile_id=?
            AND start_time >= date('now', 'start of year', 'localtime')
            GROUP BY m ORDER BY m
        """, (pid,)).fetchall()

        year_ago = (date.today() - timedelta(days=364)).isoformat()
        heatmap_data = conn.execute(
            "SELECT date(completed_at) as d, COUNT(*) as cnt FROM todos "
            "WHERE profile_id=? AND completed=1 AND completed_at>=? "
            "GROUP BY date(completed_at) ORDER BY d",
            (pid, year_ago)
        ).fetchall()
        heatmap = {row["d"]: row["cnt"] for row in heatmap_data if row["d"]}

    return render(request, "stats.html", {
        "page": "stats",
        "stats": stats,
        "chart_data": chart_data,
        "total_all": total_all,
        "total_completed": total_completed,
        "total_rate": total_rate,
        "cat_stats": cat_stats,
        "monthly_data": monthly_data,
        "monthly_events": [dict(r) for r in monthly_events],
        "heatmap": heatmap,
        "heatmap_start": year_ago,
        "heatmap_today": date.today().isoformat(),
    })


# ── Routes: Form Entry Stats (JSON) ──
@app.get("/forms/{tpl_id}/entries/stats")
async def form_entry_stats(request: Request, tpl_id: int):
    pid = get_profile_id(request)
    with get_db() as conn:
        tpl = conn.execute("SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)).fetchone()
        if not tpl:
            return JSONResponse({"dates": [], "series": []})
        fields = json.loads(tpl["fields"])
        number_fields = [f for f in fields if f.get("type") == "number"]
        entries = conn.execute(
            "SELECT entry_date, values_json FROM form_entries WHERE template_id=? AND profile_id=? "
            "AND entry_date >= date('now', '-30 days', 'localtime') ORDER BY entry_date",
            (tpl_id, pid),
        ).fetchall()
        dates = []
        series_data = {f["label"]: [] for f in number_fields}
        for e in entries:
            dates.append(e["entry_date"])
            vals = json.loads(e["values_json"])
            for f in number_fields:
                key = f["label"]
                try:
                    series_data[key].append(float(vals.get(key, 0) or 0))
                except (ValueError, TypeError):
                    series_data[key].append(0)
        series = [{"label": k, "data": v} for k, v in series_data.items()]
    return JSONResponse({"dates": dates, "series": series})


# ── Review ──
@app.get("/review", response_class=HTMLResponse)
async def review_page(request: Request):
    pid = get_profile_id(request)
    period = request.query_params.get("period", "week")
    offset = int(request.query_params.get("offset", "0"))

    today = date.today()

    if period == "month":
        first = today.replace(day=1)
        for _ in range(abs(offset)):
            first = (first - timedelta(days=1)).replace(day=1)
        if first.month == 12:
            last = first.replace(year=first.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last = first.replace(month=first.month + 1, day=1) - timedelta(days=1)
        label = first.strftime("%Y년 %m월")
    else:
        monday = today - timedelta(days=today.weekday()) + timedelta(weeks=offset)
        sunday = monday + timedelta(days=6)
        first, last = monday, sunday
        label = f"{first.strftime('%m/%d')} ~ {last.strftime('%m/%d')}"

    start_str = first.isoformat()
    end_str = last.isoformat()
    next_day = (last + timedelta(days=1)).isoformat()

    with get_db() as conn:
        total_todos = conn.execute(
            "SELECT COUNT(*) FROM todos WHERE profile_id=? AND created_at>=? AND created_at<?",
            (pid, start_str, next_day),
        ).fetchone()[0]
        completed_todos = conn.execute(
            "SELECT COUNT(*) FROM todos WHERE profile_id=? AND completed=1 AND completed_at>=? AND completed_at<?",
            (pid, start_str, next_day),
        ).fetchone()[0]

        cat_stats = conn.execute(
            "SELECT c.name, c.color, COUNT(*) as total, "
            "SUM(CASE WHEN t.completed=1 THEN 1 ELSE 0 END) as done "
            "FROM todos t LEFT JOIN categories c ON t.category_id=c.id "
            "WHERE t.profile_id=? AND t.due_date>=? AND t.due_date<=? "
            "GROUP BY t.category_id ORDER BY total DESC",
            (pid, start_str, end_str),
        ).fetchall()

        hours_data = conn.execute(
            "SELECT COALESCE(SUM(hours),0) as total_hours, COUNT(*) as log_count "
            "FROM work_logs WHERE log_date>=? AND log_date<=?",
            (start_str, end_str),
        ).fetchone()

        hours_by_cat = conn.execute(
            "SELECT c.name, c.color, SUM(w.hours) as hours "
            "FROM work_logs w LEFT JOIN categories c ON w.category_id=c.id "
            "WHERE w.log_date>=? AND w.log_date<=? "
            "GROUP BY w.category_id ORDER BY hours DESC",
            (start_str, end_str),
        ).fetchall()

        daily_completed = conn.execute(
            "SELECT date(completed_at) as d, COUNT(*) as cnt "
            "FROM todos WHERE profile_id=? AND completed=1 AND completed_at>=? AND completed_at<? "
            "GROUP BY date(completed_at) ORDER BY d",
            (pid, start_str, next_day),
        ).fetchall()

        completion_rate = round(completed_todos / total_todos * 100) if total_todos > 0 else 0

    return render(request, "review.html", {
        "page": "review",
        "period": period,
        "offset": offset,
        "label": label,
        "total_todos": total_todos,
        "completed_todos": completed_todos,
        "completion_rate": completion_rate,
        "cat_stats": [dict(r) for r in cat_stats],
        "total_hours": round(hours_data["total_hours"], 1),
        "log_count": hours_data["log_count"],
        "hours_by_cat": [dict(r) for r in hours_by_cat],
        "daily_completed": [dict(r) for r in daily_completed],
        "start_str": start_str,
        "end_str": end_str,
    })


# ── Reminders API ──
@app.get("/api/reminders")
async def get_reminders(request: Request):
    """Return upcoming todos and events that need reminders."""
    pid = get_profile_id(request)
    if not pid:
        return JSONResponse([])
    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    soon = (now + timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M")
    now_str = now.strftime("%Y-%m-%d %H:%M")
    items = []
    with get_db() as conn:
        # Overdue todos
        rows = conn.execute(
            "SELECT id, title, due_date FROM todos "
            "WHERE profile_id=? AND completed=0 AND due_date IS NOT NULL AND due_date < ? "
            "ORDER BY due_date LIMIT 10",
            (pid, today_str),
        ).fetchall()
        for r in rows:
            items.append({"type": "overdue", "id": r["id"], "title": r["title"],
                          "body": f"마감일: {r['due_date']}", "url": "/todos"})
        # Todos due today
        rows = conn.execute(
            "SELECT id, title FROM todos "
            "WHERE profile_id=? AND completed=0 AND due_date=? "
            "ORDER BY priority DESC LIMIT 10",
            (pid, today_str),
        ).fetchall()
        for r in rows:
            items.append({"type": "today", "id": r["id"], "title": r["title"],
                          "body": "오늘 마감", "url": "/todos"})
        # Events starting within 30 minutes
        rows = conn.execute(
            "SELECT id, title, start_time FROM events "
            "WHERE profile_id=? AND start_time >= ? AND start_time <= ? "
            "ORDER BY start_time LIMIT 10",
            (pid, now_str, soon),
        ).fetchall()
        for r in rows:
            items.append({"type": "event", "id": r["id"], "title": r["title"],
                          "body": f"시작: {r['start_time'][11:16]}", "url": "/calendar"})
    return JSONResponse(items)


# ── Service Worker (root scope) ──
@app.get("/sw.js")
async def service_worker():
    return FileResponse(str(BASE_DIR / "static" / "sw.js"), media_type="application/javascript")


# ── Health check ──
@app.get("/health")
async def health():
    return {"status": "ok", "app": "work-planner"}


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8002, reload=True)
