"""Excel parsing utilities for planner apps.

Handles merged cells, multi-row headers, and field-type inference.
"""

from datetime import datetime

__all__ = [
    "parse_excel_with_merges",
    "infer_field_type",
]


def parse_excel_with_merges(content: bytes):
    """Parse Excel handling merged cells and multi-row headers.

    Strategy: scan up to 20 rows to find the "real" header row -- the first row
    where >=40% of columns are filled with short text labels (not a spanning title).
    Rows above it that are part of a contiguous multi-row header block are merged
    with " > " separator. Title/metadata rows before a gap are skipped.

    Args:
        content: Raw bytes of an ``.xlsx`` file.

    Returns:
        A tuple ``(headers, sample_rows)`` where *headers* is a list of
        column-name strings and *sample_rows* is a list of row-lists (up to
        500 data rows).
    """
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # Build merged-cell value map
    merge_map = {}
    for mg in ws.merged_cells.ranges:
        val = ws.cell(row=mg.min_row, column=mg.min_col).value
        for r in range(mg.min_row, mg.max_row + 1):
            for c in range(mg.min_col, mg.max_col + 1):
                merge_map[(r, c)] = val

    def cell_val(r, c):
        if (r, c) in merge_map:
            return merge_map[(r, c)]
        return ws.cell(row=r, column=c).value

    # Detect rows dominated by wide merged cells (title rows)
    wide_merges: dict[int, int] = {}
    for mg in ws.merged_cells.ranges:
        span = mg.max_col - mg.min_col + 1
        if span >= 3:
            for r in range(mg.min_row, mg.max_row + 1):
                wide_merges.setdefault(r, 0)
                wide_merges[r] += span

    def is_tabular_header(r):
        """A tabular header has multiple distinct short text values across
        columns, not from wide merged cells (which indicate titles)."""
        if wide_merges.get(r, 0) >= max_col * 0.5:
            return False
        vals = []
        for c in range(1, max_col + 1):
            v = cell_val(r, c)
            vals.append(str(v).strip() if v is not None else "")
        distinct = set(v for v in vals if v)
        if len(distinct) < 3:
            return False
        non_empty = [v for v in vals if v]
        if len(non_empty) / max_col < 0.4:
            return False
        long_count = sum(1 for v in non_empty if len(v) > 40)
        if long_count > len(non_empty) * 0.3:
            return False
        num_count = 0
        for v in non_empty:
            try:
                float(v.replace(",", ""))
                num_count += 1
            except ValueError:
                pass
        return num_count < len(non_empty) * 0.5

    def row_fill_ratio(r):
        filled = sum(
            1
            for c in range(1, max_col + 1)
            if cell_val(r, c) is not None and str(cell_val(r, c)).strip()
        )
        return filled / max_col if max_col else 0

    # Find primary header row
    scan_limit = min(max_row + 1, 21)
    primary_header_row = None
    for r in range(1, scan_limit):
        if is_tabular_header(r):
            primary_header_row = r
            break

    if primary_header_row is None:
        for r in range(1, scan_limit):
            if row_fill_ratio(r) > 0:
                primary_header_row = r
                break
        if primary_header_row is None:
            primary_header_row = 1

    # Collect contiguous header rows above the primary
    header_rows = [primary_header_row]
    for r in range(primary_header_row - 1, 0, -1):
        if is_tabular_header(r):
            header_rows.insert(0, r)
        else:
            break

    data_start = primary_header_row + 1

    # Build column headers by joining multi-row parts with " > "
    headers = []
    for c in range(1, max_col + 1):
        parts: list[str] = []
        for r in header_rows:
            v = cell_val(r, c)
            s = str(v).strip() if v is not None else ""
            if s and s not in parts:
                parts.append(s)
        label = " > ".join(parts) if parts else f"열{c}"
        headers.append(label)

    # Drop columns that are auto-named AND have no data
    all_empty_cols = [
        i for i, h in enumerate(headers) if h.startswith("열") and h[1:].isdigit()
    ]
    if all_empty_cols:
        has_data: set[int] = set()
        check_end = min(max_row + 1, data_start + 10)
        for r in range(data_start, check_end):
            for ci in all_empty_cols:
                v = cell_val(r, ci + 1)
                if v is not None and str(v).strip():
                    has_data.add(ci)
        drop_cols = [ci for ci in all_empty_cols if ci not in has_data]
    else:
        drop_cols = []

    if drop_cols:
        headers = [h for i, h in enumerate(headers) if i not in drop_cols]
        col_indices = [
            c for c in range(1, max_col + 1) if (c - 1) not in drop_cols
        ]
    else:
        col_indices = list(range(1, max_col + 1))

    # Deduplicate header names
    seen: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in seen:
            seen[h] += 1
            headers[i] = f"{h}_{seen[h]}"
        else:
            seen[h] = 0

    # Read up to 500 data rows
    sample_rows = []
    for r in range(data_start, min(max_row + 1, data_start + 500)):
        row = []
        all_empty = True
        for c in col_indices:
            v = cell_val(r, c)
            row.append(v)
            if v is not None and str(v).strip():
                all_empty = False
        if not all_empty:
            sample_rows.append(row)

    wb.close()
    return headers, sample_rows


def infer_field_type(vals) -> str:
    """Infer the field type from a list of sample values.

    Args:
        vals: A list of cell values (may include ``None``).

    Returns:
        One of ``"text"``, ``"number"``, ``"date"``, or ``"textarea"``.
    """
    from datetime import datetime as _dt
    from datetime import time as _tm

    clean = [
        v
        for v in vals
        if v is not None and str(v).strip() and str(v).strip() != "-"
    ]
    if not clean:
        return "text"
    num_count = date_count = time_count = 0
    for v in clean[:15]:
        s = str(v).strip()
        try:
            float(s.replace(",", ""))
            num_count += 1
            continue
        except ValueError:
            pass
        if isinstance(v, _tm):
            time_count += 1
            continue
        if isinstance(v, _dt):
            if v.year == 1900 and v.month == 1 and v.day == 1:
                time_count += 1
            else:
                date_count += 1
            continue
        if hasattr(v, "strftime"):
            date_count += 1
            continue
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                datetime.strptime(s, fmt)
                date_count += 1
                break
            except ValueError:
                pass
    total = len(clean[:15])
    if time_count > total * 0.6:
        return "text"
    if num_count > total * 0.6:
        return "number"
    if date_count > total * 0.6:
        return "date"
    if any(len(str(v)) > 50 for v in clean[:15]):
        return "textarea"
    return "text"
