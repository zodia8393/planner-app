import json
import calendar
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Request, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response
from common.utils import clamp_text, fix_mojibake, validate_date_str
from common.excel import parse_excel_with_merges, infer_field_type

router = APIRouter()


def _period_range(anchor: str, freq: str):
    """Return (start_date, end_date, prev_anchor, next_anchor, label) for the period."""
    d = datetime.strptime(anchor, "%Y-%m-%d").date()
    if freq == "weekly":
        mon = d - timedelta(days=d.weekday())  # Monday
        sun = mon + timedelta(days=6)
        prev = (mon - timedelta(days=7)).isoformat()
        nxt = (mon + timedelta(days=7)).isoformat()
        label = f"{mon.month}/{mon.day}~{sun.month}/{sun.day}"
        return mon.isoformat(), sun.isoformat(), prev, nxt, label
    elif freq == "monthly":
        start = d.replace(day=1)
        last_day = calendar.monthrange(d.year, d.month)[1]
        end = d.replace(day=last_day)
        if d.month == 1:
            prev = date(d.year - 1, 12, 1).isoformat()
        else:
            prev = date(d.year, d.month - 1, 1).isoformat()
        if d.month == 12:
            nxt = date(d.year + 1, 1, 1).isoformat()
        else:
            nxt = date(d.year, d.month + 1, 1).isoformat()
        label = f"{d.year}년 {d.month}월"
        return start.isoformat(), end.isoformat(), prev, nxt, label
    elif freq == "yearly":
        start = date(d.year, 1, 1)
        end = date(d.year, 12, 31)
        prev = date(d.year - 1, 1, 1).isoformat()
        nxt = date(d.year + 1, 1, 1).isoformat()
        label = f"{d.year}년"
        return start.isoformat(), end.isoformat(), prev, nxt, label
    else:  # daily
        prev = (d - timedelta(days=1)).isoformat()
        nxt = (d + timedelta(days=1)).isoformat()
        label = anchor
        return anchor, anchor, prev, nxt, label


@router.get("/forms", response_class=HTMLResponse)
async def forms_page(request: Request):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpls = conn.execute(
            "SELECT * FROM form_templates WHERE profile_id IN (?, 0) ORDER BY profile_id ASC, updated_at DESC", (pid,)
        ).fetchall()
        entry_counts = {}
        field_counts = {}
        for t in tpls:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM form_entries WHERE template_id=? AND profile_id=?",
                (t["id"], pid),
            ).fetchone()[0]
            entry_counts[t["id"]] = cnt
            try:
                field_counts[t["id"]] = len(json.loads(t["fields"]))
            except (json.JSONDecodeError, TypeError):
                field_counts[t["id"]] = 0
    return S.render(request, "forms.html", {
        "page": "forms",
        "templates": [dict(t) for t in tpls],
        "entry_counts": entry_counts,
        "field_counts": field_counts,
    })


@router.get("/forms/new", response_class=HTMLResponse)
async def form_builder_new(request: Request):
    S = request.app.state
    S.get_profile_id(request)
    return S.render(request, "form_builder.html", {
        "page": "forms",
        "mode": "create",
        "tpl": None,
    })


@router.post("/forms/upload", response_class=HTMLResponse)
async def create_form_from_file(request: Request):
    import csv, io
    S = request.app.state
    pid = S.get_profile_id(request)
    form = await request.form()
    file = form.get("file")
    if not file or not hasattr(file, "filename") or not file.filename:
        raise HTTPException(400, "파일을 선택해주세요")

    fname = file.filename.lower()
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(413, "파일 크기 초과 (최대 10MB)")

    headers = []
    sample_rows = []

    if fname.endswith((".xlsx", ".xls")):
        headers, sample_rows = parse_excel_with_merges(content)
    elif fname.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        header_row = next(reader, None)
        if header_row:
            headers = [h.strip() if h.strip() else f"열{i+1}" for i, h in enumerate(header_row)]
        for i, row in enumerate(reader):
            if i >= 500:
                break
            sample_rows.append(row)
    else:
        raise HTTPException(400, "xlsx 또는 csv 파일만 지원합니다")

    if not headers:
        raise HTTPException(400, "헤더를 찾을 수 없습니다")

    fields = []
    for i, h in enumerate(headers):
        col_vals = [r[i] for r in sample_rows if i < len(r)]
        ftype = infer_field_type(col_vals)
        fields.append({"label": h, "type": ftype, "required": False})

    tpl_name = Path(file.filename).stem
    with S.get_db() as conn:
        cur = conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields) VALUES (?, ?, ?, ?)",
            (pid, tpl_name, f"파일 업로드로 생성 ({file.filename})", json.dumps(fields, ensure_ascii=False)),
        )
        tpl_id = cur.lastrowid

        for row in sample_rows:
            data = {}
            for j, f in enumerate(fields):
                if j < len(row) and row[j] is not None:
                    val = row[j]
                    if hasattr(val, "strftime"):
                        from datetime import datetime as _dt, time as _tm
                        if isinstance(val, _tm):
                            val = val.strftime("%H:%M")
                        elif isinstance(val, _dt):
                            if val.year == 1900 and val.month == 1 and val.day == 1:
                                val = val.strftime("%H:%M")
                            else:
                                val = val.strftime("%Y-%m-%d")
                        else:
                            val = val.strftime("%Y-%m-%d")
                    data[f["label"]] = str(val) if val else ""
                else:
                    data[f["label"]] = ""
            entry_date_val = data.get(headers[0], "") if fields[0]["type"] == "date" else date.today().isoformat()
            if not validate_date_str(entry_date_val):
                entry_date_val = date.today().isoformat()
            conn.execute(
                "INSERT INTO form_entries (template_id, profile_id, entry_date, values_json) VALUES (?, ?, ?, ?)",
                (tpl_id, pid, entry_date_val, json.dumps(data, ensure_ascii=False)),
            )

    return S.redirect(request, f"/forms/{tpl_id}/entries")


@router.post("/forms", response_class=HTMLResponse)
async def create_form_template(request: Request):
    S = request.app.state
    pid = S.get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    description = clamp_text(fix_mojibake(form.get("description", "")), 500)
    emoji = clamp_text(fix_mojibake(form.get("emoji", "")), 2) or "\U0001f4dd"
    color = form.get("color", "#6366f1") or "#6366f1"
    frequency = form.get("frequency", "daily")
    if frequency not in ("daily", "weekly", "monthly", "yearly"):
        frequency = "daily"
    if not name:
        raise HTTPException(400, "양식 이름은 필수입니다")
    fields = _parse_form_fields(form)
    if not fields:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    with S.get_db() as conn:
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color, frequency) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (pid, name, description, json.dumps(fields, ensure_ascii=False), emoji, color, frequency),
        )
    return S.redirect(request, "/forms")


@router.get("/forms/{tpl_id}/edit", response_class=HTMLResponse)
async def form_builder_edit(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
    tpl_dict = dict(tpl)
    tpl_dict["fields"] = json.loads(tpl_dict["fields"])
    tpl_dict.setdefault("frequency", "daily")
    return S.render(request, "form_builder.html", {
        "page": "forms",
        "mode": "edit",
        "tpl": tpl_dict,
    })


@router.put("/forms/{tpl_id}", response_class=HTMLResponse)
async def update_form_template(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    form = await request.form()
    name = clamp_text(fix_mojibake(form.get("name", "")), 100)
    description = clamp_text(fix_mojibake(form.get("description", "")), 500)
    emoji = clamp_text(fix_mojibake(form.get("emoji", "")), 2) or "\U0001f4dd"
    color = form.get("color", "#6366f1") or "#6366f1"
    frequency = form.get("frequency", "daily")
    if frequency not in ("daily", "weekly", "monthly", "yearly"):
        frequency = "daily"
    if not name:
        raise HTTPException(400, "양식 이름은 필수입니다")
    fields = _parse_form_fields(form)
    if not fields:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    with S.get_db() as conn:
        existing = conn.execute(
            "SELECT id FROM form_templates WHERE id=? AND profile_id=?", (tpl_id, pid)
        ).fetchone()
        if not existing:
            raise HTTPException(404)
        conn.execute(
            "UPDATE form_templates SET name=?, description=?, fields=?, emoji=?, color=?, frequency=?, updated_at=datetime('now','localtime') WHERE id=? AND profile_id=?",
            (name, description, json.dumps(fields, ensure_ascii=False), emoji, color, frequency, tpl_id, pid),
        )
    return S.redirect(request, "/forms")


@router.post("/forms/{tpl_id}/clone", response_class=HTMLResponse)
async def clone_form_template(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color, frequency) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (pid, f"{tpl['name']} (복사)", tpl["description"], tpl["fields"], tpl["emoji"], tpl["color"], tpl.get("frequency") or "daily"),
        )
    return S.redirect(request, "/forms")


@router.delete("/forms/{tpl_id}", response_class=HTMLResponse)
async def delete_form_template(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        conn.execute("DELETE FROM form_templates WHERE id=? AND profile_id=?", (tpl_id, pid))
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return S.redirect(request, "/forms")


@router.get("/forms/{form_id}/export-json")
async def export_form_json(request: Request, form_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (form_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
    fields = json.loads(tpl["fields"])
    data = {
        "name": tpl["name"],
        "description": tpl["description"] or "",
        "emoji": tpl.get("emoji") or "",
        "color": tpl.get("color") or "",
        "frequency": tpl.get("frequency") or "daily",
        "fields": fields,
    }
    content = json.dumps(data, ensure_ascii=False, indent=2)
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{tpl["name"]}.json"'},
    )


@router.post("/forms/import-json")
async def import_form_json(request: Request, file: UploadFile = File(...)):
    S = request.app.state
    pid = S.get_profile_id(request)
    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        raise HTTPException(400, "유효한 JSON 파일이 아닙니다")
    if not isinstance(data, dict) or "name" not in data or "fields" not in data:
        raise HTTPException(400, "필수 항목(name, fields)이 없습니다")
    if not isinstance(data["fields"], list) or len(data["fields"]) == 0:
        raise HTTPException(400, "최소 1개 필드가 필요합니다")
    name = clamp_text(str(data["name"]), 100)
    description = clamp_text(str(data.get("description", "")), 500)
    emoji = str(data.get("emoji", ""))[:2]
    color = str(data.get("color", "#6366f1"))[:20]
    frequency = str(data.get("frequency", "daily"))
    if frequency not in ("daily", "weekly", "monthly", "yearly"):
        frequency = "daily"
    valid_types = {"text", "textarea", "number", "dropdown", "date", "checkbox", "table"}
    fields = []
    for f in data["fields"]:
        if not isinstance(f, dict) or "label" not in f:
            continue
        field = {
            "label": clamp_text(str(f["label"]), 100),
            "type": f.get("type", "text") if f.get("type") in valid_types else "text",
            "required": bool(f.get("required", False)),
        }
        if field["type"] == "dropdown" and f.get("options"):
            field["options"] = [str(o) for o in f["options"]]
        if field["type"] == "table" and f.get("columns"):
            field["columns"] = [str(c) for c in f["columns"]]
        if f.get("default"):
            field["default"] = str(f["default"])
        if f.get("copy_prev"):
            field["copy_prev"] = True
        fields.append(field)
    if not fields:
        raise HTTPException(400, "유효한 필드가 없습니다")
    with S.get_db() as conn:
        conn.execute(
            "INSERT INTO form_templates (profile_id, name, description, fields, emoji, color, frequency) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (pid, name, description, json.dumps(fields, ensure_ascii=False), emoji, color, frequency),
        )
    return S.redirect(request, "/forms")


# ── Form Entries ──

@router.get("/forms/{tpl_id}/entries", response_class=HTMLResponse)
async def form_entries_page(request: Request, tpl_id: int, date_param: str = Query(None, alias="date")):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        tpl_dict = dict(tpl)
        tpl_dict["fields"] = json.loads(tpl_dict["fields"])
        freq = tpl.get("frequency") or "daily"
        tpl_dict["frequency"] = freq

        today_str = date_param if validate_date_str(date_param or "") else date.today().isoformat()
        period_start, period_end, prev_anchor, next_anchor, period_label = _period_range(today_str, freq)

        entries = conn.execute(
            "SELECT * FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date BETWEEN ? AND ? ORDER BY entry_date, created_at DESC",
            (tpl_id, pid, period_start, period_end),
        ).fetchall()
        parsed_entries = []
        for e in entries:
            ed = dict(e)
            ed["data"] = json.loads(ed["values_json"])
            parsed_entries.append(ed)

        # Fetch previous period's last entry for copy_prev defaults
        prev_start, prev_end, _, _, _ = _period_range(prev_anchor, freq)
        prev_entry = conn.execute(
            "SELECT values_json FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date BETWEEN ? AND ? ORDER BY entry_date DESC, created_at DESC LIMIT 1",
            (tpl_id, pid, prev_start, prev_end),
        ).fetchone()
        prev_values = json.loads(prev_entry["values_json"]) if prev_entry else {}

    # Determine if today falls within the current period
    today_iso = date.today().isoformat()
    is_today = (period_start <= today_iso <= period_end)

    field_defaults = {}
    for f in tpl_dict["fields"]:
        if f.get("copy_prev") and f["label"] in prev_values:
            field_defaults[f["label"]] = prev_values[f["label"]]
        elif f.get("default"):
            field_defaults[f["label"]] = f["default"]

    return S.render(request, "form_entries.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entries": parsed_entries,
        "current_date": today_str,
        "prev_date": prev_anchor,
        "next_date": next_anchor,
        "is_today": is_today,
        "field_defaults": field_defaults,
        "frequency": freq,
        "period_label": period_label,
        "period_start": period_start,
        "period_end": period_end,
    })


@router.get("/forms/{tpl_id}/entries/new", response_class=HTMLResponse)
async def form_entry_new(request: Request, tpl_id: int, date_param: str = Query(None, alias="date")):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        tpl_dict = dict(tpl)
        tpl_dict["fields"] = json.loads(tpl_dict["fields"])
        freq = tpl.get("frequency") or "daily"
        tpl_dict["frequency"] = freq
        entry_date = date_param if validate_date_str(date_param or "") else date.today().isoformat()

        # Fetch previous period's last entry for copy_prev defaults
        _, _, prev_anchor, _, _ = _period_range(entry_date, freq)
        prev_start, prev_end, _, _, _ = _period_range(prev_anchor, freq)
        prev_entry = conn.execute(
            "SELECT values_json FROM form_entries WHERE template_id=? AND profile_id=? AND entry_date BETWEEN ? AND ? ORDER BY entry_date DESC, created_at DESC LIMIT 1",
            (tpl_id, pid, prev_start, prev_end),
        ).fetchone()
        prev_values = json.loads(prev_entry["values_json"]) if prev_entry else {}

    field_defaults = {}
    for f in tpl_dict["fields"]:
        if f.get("copy_prev") and f["label"] in prev_values:
            field_defaults[f["label"]] = prev_values[f["label"]]
        elif f.get("default"):
            field_defaults[f["label"]] = f["default"]

    return S.render(request, "form_entry_edit.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entry": None,
        "mode": "create",
        "current_date": entry_date,
        "field_defaults": field_defaults,
        "frequency": freq,
    })


@router.post("/forms/{tpl_id}/entries", response_class=HTMLResponse)
async def create_form_entry(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    form = await request.form()
    entry_date = validate_date_str(form.get("entry_date", "")) or date.today().isoformat()
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        fields = json.loads(tpl["fields"])
        values = _collect_entry_values(form, fields)
        cur = conn.execute(
            "INSERT INTO form_entries (template_id, profile_id, entry_date, values_json) VALUES (?, ?, ?, ?)",
            (tpl_id, pid, entry_date, json.dumps(values, ensure_ascii=False)),
        )
        new_id = cur.lastrowid
    if request.headers.get("X-Inline"):
        return JSONResponse({"id": new_id})
    return S.redirect(request, f"/forms/{tpl_id}/entries?date={entry_date}")


@router.get("/forms/{tpl_id}/entries/{entry_id}/edit", response_class=HTMLResponse)
async def form_entry_edit(request: Request, tpl_id: int, entry_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        entry = conn.execute(
            "SELECT * FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        ).fetchone()
        if not entry:
            raise HTTPException(404)
    tpl_dict = dict(tpl)
    tpl_dict["fields"] = json.loads(tpl_dict["fields"])
    entry_dict = dict(entry)
    entry_dict["data"] = json.loads(entry_dict["values_json"])
    return S.render(request, "form_entry_edit.html", {
        "page": "forms",
        "tpl": tpl_dict,
        "entry": entry_dict,
        "mode": "edit",
        "field_defaults": {},
    })


@router.put("/forms/{tpl_id}/entries/{entry_id}", response_class=HTMLResponse)
async def update_form_entry(request: Request, tpl_id: int, entry_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    form = await request.form()
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        entry = conn.execute(
            "SELECT * FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        ).fetchone()
        if not entry:
            raise HTTPException(404)
        fields = json.loads(tpl["fields"])
        values = _collect_entry_values(form, fields)
        conn.execute(
            "UPDATE form_entries SET values_json=?, updated_at=datetime('now','localtime') WHERE id=? AND profile_id=?",
            (json.dumps(values, ensure_ascii=False), entry_id, pid),
        )
    if request.headers.get("X-Inline"):
        return JSONResponse({"ok": True})
    return S.redirect(request, f"/forms/{tpl_id}/entries?date={entry['entry_date']}")


@router.delete("/forms/{tpl_id}/entries/{entry_id}", response_class=HTMLResponse)
async def delete_form_entry(request: Request, tpl_id: int, entry_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        conn.execute(
            "DELETE FROM form_entries WHERE id=? AND template_id=? AND profile_id=?",
            (entry_id, tpl_id, pid),
        )
    if request.headers.get("HX-Request"):
        return HTMLResponse("")
    return S.redirect(request, f"/forms/{tpl_id}/entries")


# ── Data Export (CSV / Excel) ──

def _collect_export_data(conn, tpl_id, pid, date_filter=None):
    tpl = conn.execute("SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)).fetchone()
    if not tpl:
        return None, None, None
    fields = json.loads(tpl["fields"])
    freq = tpl.get("frequency") or "daily"
    where = "template_id=? AND profile_id=?"
    params: list = [tpl_id, pid]
    if date_filter:
        period_start, period_end, _, _, _ = _period_range(date_filter, freq)
        where += " AND entry_date BETWEEN ? AND ?"
        params.append(period_start)
        params.append(period_end)
    entries = conn.execute(f"SELECT * FROM form_entries WHERE {where} ORDER BY entry_date, id", params).fetchall()
    return tpl, fields, entries


@router.get("/forms/{tpl_id}/entries/export")
async def export_entries(request: Request, tpl_id: int, date: Optional[str] = None, fmt: str = "xlsx"):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl, fields, entries = _collect_export_data(conn, tpl_id, pid, date)
    if not tpl:
        raise HTTPException(404)

    non_table_fields = [f for f in fields if f["type"] != "table"]
    has_date_field = any(f["type"] == "date" for f in non_table_fields)
    if has_date_field:
        headers_list = [f["label"] for f in non_table_fields]
    else:
        headers_list = ["날짜"] + [f["label"] for f in non_table_fields]
    rows = []
    for e in entries:
        data = json.loads(e["values_json"])
        if has_date_field:
            row = []
        else:
            row = [e["entry_date"]]
        for f in non_table_fields:
            val = data.get(f["label"], "")
            if isinstance(val, float) and val == int(val):
                val = int(val)
            row.append(val)
        rows.append(row)

    if fmt == "csv":
        import csv, io as _io
        buf = _io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers_list)
        for r in rows:
            writer.writerow(r)
        content = buf.getvalue()
        fname = f"{tpl['name']}_{date or 'all'}.csv"
        from urllib.parse import quote
        disp = f"attachment; filename*=UTF-8''{quote(fname)}"
        return Response(content=content.encode("utf-8-sig"), media_type="text/csv",
                        headers={"Content-Disposition": disp})

    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tpl["name"][:31]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for ci, h in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border

    for ci in range(1, len(headers_list) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 15

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{tpl['name']}_{date or 'all'}.xlsx"
    from urllib.parse import quote
    disp = f"attachment; filename*=UTF-8''{quote(fname)}"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disp},
    )


@router.get("/forms/{tpl_id}/entries/stats", response_class=HTMLResponse)
async def form_entries_stats(request: Request, tpl_id: int):
    S = request.app.state
    pid = S.get_profile_id(request)
    with S.get_db() as conn:
        tpl = conn.execute(
            "SELECT * FROM form_templates WHERE id=? AND profile_id IN (?, 0)", (tpl_id, pid)
        ).fetchone()
        if not tpl:
            raise HTTPException(404)
        total = conn.execute(
            "SELECT COUNT(*) FROM form_entries WHERE template_id=? AND profile_id=?",
            (tpl_id, pid),
        ).fetchone()[0]
    return JSONResponse({"template_id": tpl_id, "total_entries": total})


# ── Helpers ──

def _parse_form_fields(form) -> list:
    """Parse field definitions from the form builder form data."""
    fields = []
    idx = 0
    while True:
        label = form.get(f"field_{idx}_label")
        if label is None:
            break
        ftype = form.get(f"field_{idx}_type", "text")
        required = form.get(f"field_{idx}_required") == "on"
        options = form.get(f"field_{idx}_options", "")
        default_val = clamp_text(fix_mojibake(form.get(f"field_{idx}_default", "")), 500)
        copy_prev = form.get(f"field_{idx}_copy_prev") == "on"
        field = {
            "label": clamp_text(fix_mojibake(label), 100),
            "type": ftype if ftype in ("text", "textarea", "number", "dropdown", "date", "checkbox", "table") else "text",
            "required": required,
            "default": default_val,
            "copy_prev": copy_prev,
        }
        if ftype == "dropdown" and options:
            field["options"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        if ftype == "table" and options:
            field["columns"] = [o.strip() for o in fix_mojibake(options).split(",") if o.strip()]
        fields.append(field)
        idx += 1
    return fields


def _collect_entry_values(form, fields: list) -> dict:
    """Collect field values from form submission."""
    values = {}
    for i, f in enumerate(fields):
        key = f"field_{i}"
        if f["type"] == "checkbox":
            values[f["label"]] = form.get(key) == "on"
        elif f["type"] == "number":
            try:
                nv = float(form.get(key, 0) or 0)
                values[f["label"]] = int(nv) if nv == int(nv) else nv
            except (ValueError, TypeError):
                values[f["label"]] = 0
        elif f["type"] == "table":
            try:
                values[f["label"]] = json.loads(form.get(key, "[]"))
            except (json.JSONDecodeError, TypeError):
                values[f["label"]] = []
        else:
            val = fix_mojibake(form.get(key, ""))
            values[f["label"]] = clamp_text(val, 5000)
    return values
